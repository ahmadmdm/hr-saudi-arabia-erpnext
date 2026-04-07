from io import BytesIO
from os.path import splitext

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cint, cstr, flt, getdate, today
from frappe.utils.file_manager import save_file
from frappe.utils.xlsxutils import make_xlsx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from saudi_hr.saudi_hr.doctype.employee_loan.employee_loan import apply_payroll_loan_deductions, get_due_loan_deduction, revert_payroll_loan_deductions
from saudi_hr.saudi_hr.utils import assert_doctype_permissions, assert_positive_basic_salary, calculate_prorated_sick_leave_deduction, get_employee_salary_components, get_gosi_rates, text_matches_tokens

GOSI_MAX_BASE = 45000.0
PREFERRED_SOURCE_WORKBOOK_SHEETS = ("كشف الرواتب طباعة", "كشف الرواتب", "كشف المصدر")
MAX_WORKBOOK_FILE_SIZE_BYTES = 10 * 1024 * 1024
ALLOWED_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
EMPLOYEE_SETUP_TEMPLATE_HEADERS = [
	"source_row",
	"payroll_employee_id",
	"payroll_employee_name",
	"national_id",
	"company",
	"department",
	"designation",
	"employee_number",
	"first_name",
	"middle_name",
	"last_name",
	"gender",
	"date_of_birth",
	"date_of_joining",
	"status",
	"remarks",
]
PAYROLL_IMPORT_TEMPLATE_HEADERS = [
	"الرقم الوظيفي",
	"الاسم",
	"مركز التكلفة",
	"مكان العمل",
	"الوظيفة",
	"بنك / كاش",
	"الاساسي",
	"بدل السكن",
	"بدل المواصلات",
	"بدلات اخرى",
	"الاجمالي",
	"ايام العمل",
	"ايام الغياب",
	"قيمة ايام الغياب",
	"ساعات التأخير",
	"قيمة التأخير",
	"خصم التأمينات",
	"إضافات",
	"خصم",
	"اجمالي الخصم",
	"صافى الراتب",
	"رقم الهوية",
	"التأمينات",
]

WORKBOOK_HEADER_ALIASES = {
	"employee_id": {"الرقم الوظيفي"},
	"employee_name": {"الاسم", "الإسم"},
	"designation": {"الوظيفة"},
	"work_location": {"مكان العمل"},
	"department": {"الإدارة"},
	"cost_center": {"مركز التكلفة", "مركز تكلفة", "cost center", "cost centre"},
	"salary_mode": {"بنك كاش", "بنك / كاش", "بنك/كاش"},
	"basic_salary": {"الاساسي"},
	"housing_allowance": {"بدل السكن"},
	"transport_allowance": {"بدل المواصلات"},
	"other_allowances": {"بدلات اخرى"},
	"gross_salary": {"الاجمالي", "اجمالي البدلات", "إجمالي البدلات"},
	"working_days": {"ايام العمل"},
	"absence_days": {"ايام الغياب", "عدد أيام الغياب"},
	"absence_value": {"قيمة ايام الغياب"},
	"late_hours": {"ساعات التأخير"},
	"late_value": {"قيمة التأخير"},
	"gosi_deduction": {"خصم التأمينات"},
	"additions": {"إضافات", "الإضافي", "اضافي", "إضافي"},
	"manual_deduction": {"خصم", "خصم السلف", "خصم السلف والاستقطاعات", "خصم الجزاءات"},
	"total_deductions": {"اجمالي الخصم", "إجمالي الخصم"},
	"net_salary": {"صافى الراتب", "صافي الراتب"},
	"national_id": {"رقم الهوية"},
	"gosi_registration": {"التأمينات", "رقم اشتراك التأمينات"},
}


class SaudiMonthlyPayroll(Document):

	def validate(self):
		self.period_label = f"{self.month} {self.year}"
		if not self.status:
			self.status = "Draft / مسودة"
		self._recalculate_employee_rows()
		self._validate_non_negative_net_salary()
		self._recalculate_totals()

	def _validate_non_negative_net_salary(self):
		for row in self.employees:
			if flt(row.net_salary) < 0:
				employee_label = row.employee_name or row.employee or _("Unknown Employee")
				frappe.throw(
					_(
						"Net salary for {0} cannot be negative. Please review loan, sick leave, and other deductions.<br>"
						"لا يمكن أن يكون صافي الراتب للموظف {0} سالباً. يرجى مراجعة خصومات القرض والإجازات والخصومات الأخرى."
					).format(employee_label),
					title=_("Negative Net Salary / صافي راتب سالب"),
				)

	def _recalculate_employee_rows(self):
		for row in self.employees:
			gross = round(
				flt(row.basic_salary)
				+ flt(row.housing_allowance)
				+ flt(row.transport_allowance)
				+ flt(row.other_allowances),
				2,
			)
			total_deductions = round(
				flt(row.gosi_employee_deduction)
				+ flt(row.sick_leave_deduction)
				+ flt(row.loan_deduction)
				+ flt(getattr(row, "other_deductions", 0.0)),
				2,
			)
			row.gross_salary = gross
			row.total_deductions = total_deductions
			row.net_salary = round(gross + flt(row.overtime_addition) - total_deductions, 2)
			row.cost_center = _resolve_cost_center_link(getattr(row, "cost_center", None), self.company) or getattr(row, "cost_center", None)

	def _recalculate_totals(self):
		"""إعادة حساب الإجماليات من الجدول الفرعي."""
		self.total_employees = len(self.employees)
		self.total_gross = round(sum(flt(r.gross_salary) for r in self.employees), 2)
		self.total_gosi_deductions = round(sum(flt(r.gosi_employee_deduction) for r in self.employees), 2)
		self.total_sick_deductions = round(sum(flt(r.sick_leave_deduction) for r in self.employees), 2)
		self.total_loan_deductions = round(sum(flt(r.loan_deduction) for r in self.employees), 2)
		self.total_other_deductions = round(sum(flt(getattr(r, "other_deductions", 0.0)) for r in self.employees), 2)
		self.total_overtime = round(sum(flt(r.overtime_addition) for r in self.employees), 2)
		self.total_net_payable = round(sum(flt(r.net_salary) for r in self.employees), 2)

	def on_submit(self):
		apply_payroll_loan_deductions(self)
		self.db_set("status", "Completed / مكتمل")

	def on_cancel(self):
		revert_payroll_loan_deductions(self)
		self.db_set("status", "Cancelled / ملغى")


# ─── Whitelist API Methods ───────────────────────────────────────────────────────

@frappe.whitelist()
def fetch_employees(doc_name: str):
	"""
	جلب جميع الموظفين النشطين للشركة وملء الجدول الفرعي.
	يُستدعى من زر JavaScript.
	"""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "write", doc=doc, throw=True)

	employees = frappe.get_all(
		"Employee",
		filters={"company": doc.company, "status": "Active"},
		fields=_get_employee_fetch_fields(),
		order_by="employee_name",
	)

	if not employees:
		doc.set("employees", [])
		doc._recalculate_totals()
		doc.save()
		return {"count": 0, "total_net": 0.0}

	emp_names = [e["name"] for e in employees]

	# ── جلب بدلات العقود ورواتبها دفعةً واحدة ───────────────────────────────
	contract_rows = frappe.db.sql(
		"""SELECT employee, basic_salary, housing_allowance, transport_allowance, other_allowances, total_salary
		   FROM `tabSaudi Employment Contract`
		   WHERE employee IN %(names)s
		     AND contract_status = 'Active / نشط'
		     AND docstatus = 1
		   ORDER BY start_date DESC""",
		{"names": emp_names},
		as_dict=True,
	)
	contract_map = {}
	for row in contract_rows:
		if row["employee"] not in contract_map:
			contract_map[row["employee"]] = row

	# ── حساب نطاق الشهر مرة واحدة ──────────────────────────────────────────
	import calendar as _cal
	month_num = _month_name_to_num(doc.month)
	month_start = f"{doc.year}-{month_num:02d}-01"
	last_day = _cal.monthrange(int(doc.year), month_num)[1]
	month_end = f"{doc.year}-{month_num:02d}-{last_day:02d}"

	# ── جلب الإجازات المرضية دفعةً واحدة ───────────────────────────────────
	sick_rows = frappe.db.sql(
		"""SELECT employee, leave_pay_amount, total_days, daily_salary, from_date, to_date
		   FROM `tabSaudi Sick Leave`
		   WHERE employee IN %(names)s AND docstatus=1
		     AND from_date <= %(end)s AND to_date >= %(start)s""",
		{"names": emp_names, "start": month_start, "end": month_end},
		as_dict=True,
	)
	sick_map: dict = {}
	for sr in sick_rows:
		sick_map.setdefault(sr["employee"], []).append(sr)

	# ── جلب العمل الإضافي المعتمد دفعةً واحدة ───────────────────────────────
	ot_rows = frappe.db.sql(
		"""SELECT employee, overtime_amount, approval_status
		   FROM `tabOvertime Request`
		   WHERE employee IN %(names)s AND docstatus=1
		     AND date BETWEEN %(start)s AND %(end)s""",
		{"names": emp_names, "start": month_start, "end": month_end},
		as_dict=True,
	)
	ot_map: dict = {}
	for ot in ot_rows:
		if not text_matches_tokens(ot.get("approval_status"), "approved", "موافق"):
			continue
		ot_map[ot["employee"]] = ot_map.get(ot["employee"], 0.0) + flt(ot["overtime_amount"])

	loan_map = {emp_name: get_due_loan_deduction(emp_name, month_num, int(doc.year)) for emp_name in emp_names}

	# مسح الجدول الحالي
	doc.set("employees", [])

	for emp in employees:
		contract = contract_map.get(emp["name"])
		basic = flt(contract["basic_salary"]) if contract else 0.0
		if not basic:
			basic = get_employee_salary_components(emp["name"])["basic_salary"]
		assert_positive_basic_salary(emp.get("employee_name") or emp["name"], basic, _("fetching payroll / جلب الرواتب"))
		housing = flt(contract["housing_allowance"]) if contract else 0.0
		transport = flt(contract["transport_allowance"]) if contract else 0.0
		other = flt(contract["other_allowances"]) if contract else 0.0
		gross = round(basic + housing + transport + other, 2)

		gosi_rate = flt(get_gosi_rates(emp.get("nationality") or "").get("employee_rate"))
		gosi_base = min(basic, GOSI_MAX_BASE)
		gosi_deduction = round(gosi_base * gosi_rate / 100, 2)

		sick_deduction = calculate_prorated_sick_leave_deduction(
			sick_map.get(emp["name"], []),
			month_start,
			month_end,
			round(basic / 30, 2),
		)

		overtime = round(ot_map.get(emp["name"], 0.0), 2)
		loan_info = loan_map.get(emp["name"], {"loan_deduction": 0.0, "installment_names": []})
		loan_deduction = round(flt(loan_info.get("loan_deduction")), 2)
		total_deductions = round(gosi_deduction + sick_deduction + loan_deduction, 2)
		net = round(gross + overtime - total_deductions, 2)

		doc.append("employees", {
			"employee": emp["name"],
			"employee_name": emp.get("employee_name", ""),
			"department": emp.get("department", ""),
			"nationality": emp.get("nationality", ""),
			"work_location": "",
			"designation": "",
			"salary_mode": "",
			"gosi_registration": "",
			"working_days": 0.0,
			"absence_days": 0.0,
			"late_hours": 0.0,
			"basic_salary": basic,
			"housing_allowance": housing,
			"transport_allowance": transport,
			"other_allowances": other,
			"gross_salary": gross,
			"gosi_employee_deduction": gosi_deduction,
			"sick_leave_deduction": round(sick_deduction, 2),
			"loan_deduction": loan_deduction,
			"total_deductions": total_deductions,
			"overtime_addition": overtime,
			"loan_installments": ", ".join(loan_info.get("installment_names", [])),
			"net_salary": net,
				"cost_center": _get_company_default_cost_center(emp.get("company") or ""),
		})

	doc._recalculate_totals()
	doc.save()
	return {"count": len(employees), "total_net": doc.total_net_payable}


@frappe.whitelist()
def calculate_employee_row(employee: str, month: str, year: int):
	"""
	حساب الراتب الشهري لموظف واحد وإرجاع dict للتحديث في الواجهة.
	"""
	emp_doc = frappe.get_all(
		"Employee",
		filters={"name": employee},
		fields=_get_employee_fetch_fields(),
		limit=1,
	)
	if not emp_doc:
		frappe.throw(_("Employee not found"))
	row = _build_employee_row(emp_doc[0], month, int(year))
	return row


@frappe.whitelist()
def import_payroll_workbook(doc_name: str, file_url: str | None = None):
	"""استيراد ملف رواتب خارجي إلى كشف الرواتب الشهري."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "write", doc=doc, throw=True)

	workbook_url = file_url or doc.source_workbook
	if not workbook_url:
		frappe.throw(_("Attach the source payroll workbook first.<br>أرفق ملف الرواتب المصدر أولاً."))

	analysis = _preview_payroll_workbook(doc.company, workbook_url)
	import_rows = analysis["import_rows"]
	warnings = analysis["warnings"]
	validation_summary = _validate_payroll_workbook_rows(doc.company, _extract_source_workbook_rows(_get_attached_file_content(workbook_url)))
	if validation_summary["error_count"]:
		sample_errors = "<br>".join(validation_summary["errors"][:10])
		extra_errors = max(validation_summary["error_count"] - 10, 0)
		extra_text = _("<br>ويوجد {0} أخطاء إضافية.").format(extra_errors) if extra_errors else ""
		frappe.throw(
			_(
				"لا يمكن استيراد ملف الرواتب قبل معالجة أخطاء التحقق.<br>{0}{1}"
			).format(sample_errors, extra_text),
			title=_("أخطاء مانعة في ملف الرواتب / Blocking Payroll Workbook Errors"),
		)

	if not import_rows:
		frappe.throw(
			_(
				"Could not import any payroll rows from the workbook. "
				f"Employees in company: {analysis['company_employee_count']}. "
				f"Unmatched workbook rows: {analysis['unmatched_rows']}. "
				f"Sample IDs: {', '.join(analysis['sample_unmatched'][:5]) or 'N/A'}"
				"<br>تعذّر استيراد أي صفوف رواتب من الملف بسبب عدم وجود مطابقة مع بيانات الموظفين الحالية."
			),
			title=_("No Importable Rows / لا توجد صفوف قابلة للاستيراد"),
		)

	doc.source_workbook = workbook_url
	doc.set("employees", [])
	for row in import_rows:
		doc.append("employees", row)

	auto_create_result = _auto_create_missing_employees_for_import(doc)

	doc._recalculate_employee_rows()
	doc._recalculate_totals()
	doc.save()
	_add_payroll_audit_comment(
		doc,
		_("Imported {0} payroll rows from workbook {1}. Warnings: {2}. Auto-created employees: {3}. Remaining unlinked rows: {4}.").format(
			len(import_rows),
			workbook_url,
			len(warnings),
			auto_create_result["created_count"],
			auto_create_result["remaining_unlinked_rows"],
		),
	)

	return {
		"count": len(import_rows),
		"warnings": warnings,
		"total_net": doc.total_net_payable,
		"auto_create_enabled": auto_create_result["enabled"],
		"created_count": auto_create_result["created_count"],
		"linked_count": auto_create_result["linked_count"],
		"skipped": auto_create_result["skipped"],
		"remaining_unlinked_rows": auto_create_result["remaining_unlinked_rows"],
	}


@frappe.whitelist()
def preview_payroll_workbook_import(doc_name: str, file_url: str | None = None):
	"""معاينة ملف الرواتب قبل الاستيراد وإظهار نسبة المطابقة."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "read", doc=doc, throw=True)
	workbook_url = file_url or doc.source_workbook
	if not workbook_url:
		frappe.throw(_("Attach the source payroll workbook first.<br>أرفق ملف الرواتب المصدر أولاً."))
	return _preview_payroll_workbook(doc.company, workbook_url)


@frappe.whitelist()
def validate_payroll_workbook(doc_name: str, file_url: str | None = None):
	"""التحقق من ملف الرواتب قبل الاستيراد مع إظهار الأخطاء والتحذيرات."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "read", doc=doc, throw=True)
	workbook_url = file_url or doc.source_workbook
	if not workbook_url:
		frappe.throw(_("Attach the source payroll workbook first.<br>أرفق ملف الرواتب المصدر أولاً."))

	raw_rows = _extract_source_workbook_rows(_get_attached_file_content(workbook_url))
	return _validate_payroll_workbook_rows(doc.company, raw_rows)


@frappe.whitelist()
def download_payroll_workbook_gap_report(doc_name: str, file_url: str | None = None):
	"""إنشاء ملف Excel لصفوف الرواتب غير المطابقة مع بيانات الموظفين الحالية."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "read", doc=doc, throw=True)
	workbook_url = file_url or doc.source_workbook
	if not workbook_url:
		frappe.throw(_("Attach the source payroll workbook first.<br>أرفق ملف الرواتب المصدر أولاً."))

	summary = _preview_payroll_workbook(doc.company, workbook_url)
	gap_rows = _build_gap_report_rows(summary.get("unmatched_details", []))
	if len(gap_rows) == 1:
		frappe.throw(
			_("No unmatched workbook rows were found.<br>لا توجد صفوف غير مطابقة في ملف الرواتب."),
			title=_("No Gaps / لا توجد فجوات"),
		)

	file_name = f"payroll-workbook-gaps-{doc.name}.xlsx"
	file_doc = save_file(file_name, make_xlsx(gap_rows, "Payroll Gaps").getvalue(), doc.doctype, doc.name, is_private=1)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name, "row_count": len(gap_rows) - 1}


@frappe.whitelist()
def download_payroll_import_template(doc_name: str):
	"""إنشاء قالب Excel فارغ لرفع الرواتب دفعة واحدة مع تعليمات مبسطة."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "read", doc=doc, throw=True)

	file_name = f"payroll-import-template-{doc.name}-{frappe.generate_hash(length=6)}.xlsx"
	file_doc = save_file(
		file_name,
		_build_payroll_import_template_workbook(doc).getvalue(),
		doc.doctype,
		doc.name,
		is_private=1,
	)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name}


@frappe.whitelist()
def download_employee_setup_template(doc_name: str, file_url: str | None = None):
	"""إنشاء قالب إعداد موظفين من صفوف الرواتب غير المطابقة."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "read", doc=doc, throw=True)
	workbook_url = file_url or doc.source_workbook
	if not workbook_url:
		frappe.throw(_("Attach the source payroll workbook first.<br>أرفق ملف الرواتب المصدر أولاً."))

	summary = _preview_payroll_workbook(doc.company, workbook_url)
	template_rows = _build_employee_setup_template_rows(doc.company, summary.get("unmatched_details", []))
	if len(template_rows) == 1:
		frappe.throw(
			_("No unmatched workbook rows were found.<br>لا توجد صفوف غير مطابقة في ملف الرواتب."),
			title=_("No Gaps / لا توجد فجوات"),
		)

	file_name = f"employee-setup-template-{doc.name}.xlsx"
	file_doc = save_file(file_name, make_xlsx(template_rows, "Employee Setup").getvalue(), doc.doctype, doc.name, is_private=1)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name, "row_count": len(template_rows) - 1}


@frappe.whitelist()
def import_employee_setup_workbook(doc_name: str, file_url: str | None = None):
	"""استيراد ملف إعداد الموظفين المكتمل وإنشاء سجلات Employee الناقصة."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "write", doc=doc, throw=True)
	workbook_url = file_url or doc.employee_setup_workbook
	if not workbook_url:
		frappe.throw(_("Attach the completed employee setup workbook first.<br>أرفق ملف إعداد الموظفين المكتمل أولاً."))

	doc.db_set("employee_setup_workbook", workbook_url)
	rows = _extract_employee_setup_rows(_get_attached_file_content(workbook_url))
	created, skipped = _import_employee_setup_rows(doc.company, rows)
	_add_payroll_audit_comment(
		doc,
		_("Imported employee setup workbook {0}. Created employees: {1}. Skipped rows: {2}.").format(
			workbook_url,
			len(created),
			len(skipped),
		),
	)
	return {"created_count": len(created), "skipped": skipped, "created_employees": created}


@frappe.whitelist()
def autofill_employee_setup_workbook_names(doc_name: str, file_url: str | None = None):
	"""تعبئة حقول الاسم تلقائياً في ملف إعداد الموظفين من اسم الموظف القادم من ملف الرواتب."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "write", doc=doc, throw=True)
	workbook_url = file_url or doc.employee_setup_workbook
	if not workbook_url:
		frappe.throw(_("Attach the employee setup workbook first.<br>أرفق ملف إعداد الموظفين أولاً."))

	rows = _extract_employee_setup_rows(_get_attached_file_content(workbook_url))
	autofilled_rows, updated_count = _autofill_employee_setup_names(rows)
	file_name = f"employee-setup-autofilled-{doc.name}.xlsx"
	file_doc = save_file(
		file_name,
		make_xlsx(_build_employee_setup_workbook_rows(autofilled_rows), "Employee Setup").getvalue(),
		doc.doctype,
		doc.name,
		is_private=1,
	)
	doc.db_set("employee_setup_workbook", file_doc.file_url)
	return {
		"file_url": file_doc.file_url,
		"file_name": file_doc.file_name,
		"updated_count": updated_count,
	}


@frappe.whitelist()
def create_basic_employees_from_payroll(
	doc_name: str,
	default_gender: str,
	default_date_of_birth,
	default_date_of_joining=None,
	default_status: str = "Active",
):
	"""Create placeholder Employee records for imported payroll rows that are not linked yet."""
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "write", doc=doc, throw=True)
	defaults = _normalize_basic_employee_creation_defaults(
		default_gender,
		default_date_of_birth,
		default_date_of_joining or doc.posting_date,
		default_status,
	)
	assert_doctype_permissions("Employee", "create")
	created, linked, skipped = _create_basic_employees_for_payroll(doc, defaults)
	doc.save()
	_add_payroll_audit_comment(
		doc,
		_("Created {0} placeholder employees from payroll. Linked existing rows: {1}. Skipped rows: {2}.").format(
			len(created),
			linked,
			len(skipped),
		),
	)

	remaining_unlinked = sum(1 for row in doc.employees if not cstr(row.employee or "").strip())
	return {
		"created_count": len(created),
		"linked_count": linked,
		"created_employees": created,
		"skipped": skipped,
		"remaining_unlinked_rows": remaining_unlinked,
	}


def _auto_create_missing_employees_for_import(doc) -> dict:
	remaining_unlinked = sum(1 for row in doc.employees if not cstr(row.employee or "").strip())
	if not cint(getattr(doc, "auto_create_missing_employees", 0)):
		return {
			"enabled": False,
			"created_count": 0,
			"linked_count": 0,
			"skipped": [],
			"remaining_unlinked_rows": remaining_unlinked,
		}

	try:
		assert_doctype_permissions("Employee", "create")
	except frappe.PermissionError:
		return {
			"enabled": True,
			"created_count": 0,
			"linked_count": 0,
			"skipped": [
				_(
					"Automatic employee creation is enabled, but you do not have permission to create Employee records. "
					"The payroll rows were imported and remain unlinked until an authorized user creates the employee masters.<br>"
					"تفعيل إنشاء الموظفين التلقائي موجود، لكن لا تملك صلاحية إنشاء سجلات Employee. "
					"تم استيراد صفوف الرواتب وستبقى غير مرتبطة حتى يقوم مستخدم مخول بإنشاء سجلات الموظفين الأساسية."
				)
			],
			"remaining_unlinked_rows": remaining_unlinked,
		}

	defaults = _normalize_basic_employee_creation_defaults(
		cstr(getattr(doc, "auto_create_default_gender", None) or "Prefer not to say").strip() or "Prefer not to say",
		getattr(doc, "auto_create_default_date_of_birth", None) or "1990-01-01",
		getattr(doc, "auto_create_default_date_of_joining", None) or doc.posting_date,
		"Active",
	)
	created, linked, skipped = _create_basic_employees_for_payroll(doc, defaults)
	remaining_unlinked = sum(1 for row in doc.employees if not cstr(row.employee or "").strip())
	return {
		"enabled": True,
		"created_count": len(created),
		"linked_count": linked,
		"skipped": skipped,
		"remaining_unlinked_rows": remaining_unlinked,
	}


@frappe.whitelist()
def create_journal_entry_from_payroll(doc_name: str):
	"""
	إنشاء قيد يومي مباشر من بيانات Saudi Monthly Payroll
	بدلاً من الاعتماد على Payroll Entry في ERPNext/HRMS.
	يُستدعى من زر "Create Journal Entry".
	"""
	import calendar as _cal
	doc = frappe.get_doc("Saudi Monthly Payroll", doc_name)
	frappe.has_permission("Saudi Monthly Payroll", "write", doc=doc, throw=True)

	if doc.payroll_journal_entry:
		frappe.throw(
			_(f"Journal Entry already created: {doc.payroll_journal_entry}<br>"
			  f"القيد اليومي موجود بالفعل: {doc.payroll_journal_entry}"),
			title=_("Already Created / موجودة مسبقاً"),
		)

	if not doc.employees:
		frappe.throw(
			_("No employees in the payroll. Please fetch employees first.<br>"
			  "لا يوجد موظفون. الرجاء جلب الموظفين أولاً."),
			title=_("No Employees / لا يوجد موظفون"),
		)

	company = doc.company
	month_num = _month_name_to_num(doc.month)
	last_day = _cal.monthrange(int(doc.year), month_num)[1]
	posting_date = f"{doc.year}-{month_num:02d}-{last_day:02d}"

	# ── حسابات الإجمالي ─────────────────────────────────────────────────────
	total_gross = flt(doc.total_gross)
	total_gosi = flt(doc.total_gosi_deductions)
	total_sick = flt(doc.total_sick_deductions)
	total_loan = flt(doc.total_loan_deductions)
	total_other = flt(getattr(doc, "total_other_deductions", 0.0))
	total_overtime = flt(doc.total_overtime)
	total_net = flt(doc.total_net_payable)
	total_salary_cost = round(total_gross + total_overtime - total_sick - total_other, 2)

	# ── الحسابات المحاسبية ───────────────────────────────────────────────────
	company_accounts = frappe.get_all(
		"Account",
		filters={"company": company, "is_group": 0},
		fields=["name", "account_name", "account_type", "root_type"],
	)
	expense_account = _find_company_account(company_accounts, root_type="Expense", name_terms=("salary",)) or _find_company_account(company_accounts, root_type="Expense")
	payable_account = _get_payroll_payable_account(company, accounts=company_accounts)
	gosi_payable_account = _find_company_account(company_accounts, root_type="Liability", name_terms=("gosi",)) or payable_account
	loan_receivable_account = _find_company_account(company_accounts, root_type="Asset", name_terms=("loan",)) or _find_company_account(company_accounts, account_type="Receivable")

	if not expense_account or not payable_account:
		frappe.throw(
			_("Could not find Salary expense/payable accounts. "
			  "Please configure them in the Chart of Accounts.<br>"
			  "تعذّر إيجاد حسابات الرواتب. يرجى إعداد شجرة الحسابات."),
			title=_("Account Not Found / حساب غير موجود"),
		)

	# ── بناء قيود الحساب ────────────────────────────────────────────────────
	# Dr. Salary Expense    → إجمالي الرواتب
	# Cr. GOSI Payable      → اشتراكات GOSI للموظف
	# Cr. Salary Payable    → صافي المستحق للموظفين
	accounts = [
	]
	company_cost_center = _get_company_default_cost_center(company)
	accounts.extend(_build_payroll_expense_account_rows(doc, expense_account, company_cost_center))
	if total_gosi > 0:
		accounts.append({
			"account": gosi_payable_account,
			"credit_in_account_currency": total_gosi,
			"cost_center": company_cost_center,
		})
	if total_loan > 0 and loan_receivable_account:
		accounts.append({
			"account": loan_receivable_account,
			"credit_in_account_currency": total_loan,
			"cost_center": company_cost_center,
		})
	accounts.append({
		"account": payable_account,
		"credit_in_account_currency": total_net,
		"cost_center": company_cost_center,
	})
	accounts = _merge_journal_entry_accounts(accounts)

	je = frappe.get_doc({
		"doctype": "Journal Entry",
		"voucher_type": "Journal Entry",
		"company": company,
		"posting_date": posting_date,
		"user_remark": (
			f"Monthly Payroll — {doc.month} {doc.year} — "
			f"{doc.total_employees} employees — Net: {total_net:,.2f} SAR"
		),
		"accounts": accounts,
	})
	assert_doctype_permissions("Journal Entry", ("create", "submit"))
	je.insert()
	je.submit()

	doc.db_set("payroll_journal_entry", je.name)
	doc.db_set("status", "Completed / مكتمل")

	frappe.msgprint(
		_(f"Journal Entry <b>{je.name}</b> created for {doc.month} {doc.year} payroll.<br>"
		  f"تم إنشاء القيد اليومي <b>{je.name}</b> لرواتب {doc.month} {doc.year}."),
		title=_("Journal Entry Created / تم إنشاء القيد"),
		indicator="green",
	)
	return je.name


def _get_payroll_payable_account(company: str, accounts: list[dict] | None = None) -> str | None:
	accounts = accounts or frappe.get_all(
		"Account",
		filters={"company": company, "root_type": "Liability", "is_group": 0},
		fields=["name", "account_name", "account_type", "root_type"],
	)
	accounts = [
		account for account in accounts
		if cstr(account.get("root_type") or "liability").strip().lower() == "liability"
	]

	def rank(account: dict):
		name = cstr(account.get("account_name") or account.get("name") or "").strip().lower()
		account_type = cstr(account.get("account_type") or "").strip().lower()

		if "payroll payable" in name:
			return (0, name)
		if "salary payable" in name:
			return (1, name)
		if "payroll" in name and account_type != "payable":
			return (2, name)
		if "salary" in name and account_type != "payable":
			return (3, name)
		if "payable" in name and account_type != "payable":
			return (4, name)
		return (99, name)

	preferred = [account for account in accounts if rank(account)[0] < 99]
	if preferred:
		preferred.sort(key=rank)
		return preferred[0]["name"]

	payable_accounts = [account for account in accounts if cstr(account.get("account_type") or "").strip().lower() == "payable"]
	if payable_accounts:
		frappe.throw(
			_(
				"Could not find a payroll payable liability account. Generic payable accounts require Party details. "
				"Please configure an account such as Payroll Payable or Salary Payable.<br>"
				"تعذّر العثور على حساب التزامات مناسب للرواتب. حسابات الدائنين العامة تتطلب طرفاً محاسبياً، لذا يرجى إعداد حساب مثل Payroll Payable أو Salary Payable."
			),
			title=_("Payroll Payable Account Missing / حساب رواتب دائن غير موجود"),
		)

	return None


def _find_company_account(accounts: list[dict], root_type: str | None = None, account_type: str | None = None, name_terms: tuple[str, ...] = ()) -> str | None:
	for account in accounts:
		account_root_type = cstr(account.get("root_type") or "").strip().lower()
		account_account_type = cstr(account.get("account_type") or "").strip().lower()
		account_name = cstr(account.get("account_name") or account.get("name") or "").strip().lower()
		if root_type and account_root_type != root_type.lower():
			continue
		if account_type and account_account_type != account_type.lower():
			continue
		if name_terms and not any(term.lower() in account_name for term in name_terms):
			continue
		return account.get("name")
	return None


def _get_company_default_cost_center(company: str) -> str:
	company_name = cstr(company or "").strip()
	if not company_name:
		return ""

	default_cost_center = cstr(frappe.db.get_value("Company", company_name, "cost_center") or "").strip()
	resolved_default = _get_postable_cost_center(default_cost_center)
	if resolved_default:
		return resolved_default

	company_cost_centers = frappe.get_all(
		"Cost Center",
		filters={"company": company_name, "is_group": 0},
		fields=["name"],
		order_by="lft asc",
		limit=1,
	)
	return company_cost_centers[0]["name"] if company_cost_centers else ""


def _get_company_root_cost_center(company: str) -> str:
	company_name = cstr(company or "").strip()
	if not company_name:
		return ""

	root_row = frappe.get_all(
		"Cost Center",
		filters={"company": company_name, "is_group": 1},
		fields=["name"],
		order_by="lft asc",
		limit=1,
	)
	return root_row[0]["name"] if root_row else ""


def _get_postable_cost_center(cost_center: str, company: str = "") -> str:
	name = cstr(cost_center or "").strip()
	if not name:
		return ""
	if not frappe.db.exists("Cost Center", name) and company:
		name = cstr(frappe.db.get_value("Cost Center", {"company": company, "cost_center_name": name}, "name") or "").strip()
	if not name or not frappe.db.exists("Cost Center", name):
		return ""

	row = frappe.db.get_value("Cost Center", name, ["is_group", "lft", "rgt"], as_dict=True)
	if not row:
		return ""
	if not cint(row.is_group):
		return name

	child = frappe.get_all(
		"Cost Center",
		filters={
			"is_group": 0,
			"lft": [">", cint(row.lft or 0)],
			"rgt": ["<", cint(row.rgt or 0)],
		},
		fields=["name"],
		order_by="lft asc",
		limit=1,
	)
	return child[0]["name"] if child else ""


def _create_company_cost_center(company: str, cost_center_label: str) -> str:
	company_name = cstr(company or "").strip()
	label = cstr(cost_center_label or "").strip()
	if not company_name or not label:
		return ""

	existing = _get_postable_cost_center(label, company_name)
	if existing:
		return existing

	parent_cost_center = _get_company_root_cost_center(company_name)
	if not parent_cost_center:
		return ""

	cost_center = frappe.get_doc({
		"doctype": "Cost Center",
		"cost_center_name": label,
		"company": company_name,
		"parent_cost_center": parent_cost_center,
		"is_group": 0,
	})
	cost_center.insert(ignore_permissions=True)
	return cost_center.name


def _resolve_cost_center_link(value, company: str) -> str:
	postable_cost_center = _get_postable_cost_center(value, company)
	if postable_cost_center:
		return postable_cost_center

	created_cost_center = _create_company_cost_center(company, value)
	if created_cost_center:
		return created_cost_center

	return _get_company_default_cost_center(company)


def _build_payroll_expense_account_rows(doc, expense_account: str, default_cost_center: str) -> list[dict]:
	grouped_amounts = {}
	ordered_cost_centers = []

	for row in doc.employees:
		row_salary_cost = round(
			flt(row.gross_salary)
			+ flt(row.overtime_addition)
			- flt(row.sick_leave_deduction)
			- flt(getattr(row, "other_deductions", 0.0)),
			2,
		)
		if not row_salary_cost:
			continue

		cost_center = _resolve_cost_center_link(getattr(row, "cost_center", None), doc.company) or default_cost_center
		if cost_center not in grouped_amounts:
			grouped_amounts[cost_center] = 0.0
			ordered_cost_centers.append(cost_center)
		grouped_amounts[cost_center] += row_salary_cost

	accounts = []
	for cost_center in ordered_cost_centers:
		accounts.append({
			"account": expense_account,
			"debit_in_account_currency": round(grouped_amounts[cost_center], 2),
			"cost_center": cost_center,
		})

	if not accounts:
		accounts.append({
			"account": expense_account,
			"debit_in_account_currency": 0.0,
			"cost_center": default_cost_center,
		})

	return accounts


def _build_payroll_import_template_workbook(doc) -> BytesIO:
	workbook = Workbook()
	instructions_sheet = workbook.active
	instructions_sheet.title = "Instructions"
	instructions_sheet.sheet_view.rightToLeft = True
	example_sheet = workbook.create_sheet("Example")
	example_sheet.sheet_view.rightToLeft = True
	payload_sheet = workbook.create_sheet(PREFERRED_SOURCE_WORKBOOK_SHEETS[0])
	payload_sheet.sheet_view.rightToLeft = True

	title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
	header_fill = PatternFill(fill_type="solid", fgColor="EAF2E3")
	bold_font = Font(bold=True)
	wrapped_alignment = Alignment(wrap_text=True, vertical="top")

	instruction_rows = [
		["قالب رفع الرواتب", "استخدم هذا الملف لرفع الرواتب دفعة واحدة من شاشة مسير الرواتب الشهري السعودي."],
		["1", "املأ البيانات داخل ورقة كشف الرواتب فقط ولا تغير أسماء الأعمدة في الصف الأول."],
		["2", "الحد الأدنى المطلوب لكل صف عادة: الرقم الوظيفي، الاسم، مركز التكلفة، الأساسي، الإجمالي، إجمالي الخصم، صافي الراتب."],
		["3", "إذا كان راتب الموظف موزعاً على أكثر من مركز تكلفة، كرر نفس الموظف في أكثر من صف مع تغيير مركز التكلفة والقيم المالية لكل صف."],
		["4", "صيغة مركز التكلفة في هذا القالب هي مركز التكلفة مباشرة. إذا رفعت ملفاً قديماً يستخدم الإدارة فسيتم التعامل معه أيضاً كمركز تكلفة."],
		["5", f"اسم الشيت المطلوب للاستيراد: {PREFERRED_SOURCE_WORKBOOK_SHEETS[0]}"],
		["6", "بعد تعبئة الملف، أرفقه في حقل ملف الرواتب المصدر ثم اضغط استيراد ملف الرواتب."],
	]

	for row_index, row in enumerate(instruction_rows, start=1):
		for column_index, value in enumerate(row, start=1):
			cell = instructions_sheet.cell(row=row_index, column=column_index, value=value)
			cell.alignment = wrapped_alignment
			if row_index == 1:
				cell.font = bold_font
				cell.fill = title_fill

	instructions_sheet.column_dimensions["A"].width = 18
	instructions_sheet.column_dimensions["B"].width = 110

	example_rows = _build_payroll_import_example_rows(doc.company)
	for column_index, header in enumerate(PAYROLL_IMPORT_TEMPLATE_HEADERS, start=1):
		example_cell = example_sheet.cell(row=1, column=column_index, value=header)
		example_cell.font = bold_font
		example_cell.fill = header_fill
		example_cell.alignment = Alignment(horizontal="center", vertical="center")
		example_sheet.column_dimensions[example_cell.column_letter].width = 18

		cell = payload_sheet.cell(row=1, column=column_index, value=header)
		cell.font = bold_font
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal="center", vertical="center")
		payload_sheet.column_dimensions[cell.column_letter].width = 18

	for row_index, row_values in enumerate(example_rows, start=2):
		for column_index, value in enumerate(row_values, start=1):
			example_sheet.cell(row=row_index, column=column_index, value=value)

	example_sheet.freeze_panes = "A2"
	example_sheet.auto_filter.ref = f"A1:{example_sheet.cell(row=1, column=len(PAYROLL_IMPORT_TEMPLATE_HEADERS)).column_letter}{len(example_rows) + 1}"
	payload_sheet.freeze_panes = "A2"
	payload_sheet.auto_filter.ref = f"A1:{payload_sheet.cell(row=1, column=len(PAYROLL_IMPORT_TEMPLATE_HEADERS)).column_letter}1"

	output = BytesIO()
	workbook.save(output)
	output.seek(0)
	return output


def _build_payroll_import_example_rows(company: str) -> list[list]:
	default_cost_center = _get_company_default_cost_center(company) or "Main"
	second_cost_center = _get_second_postable_cost_center(company, default_cost_center) or "فرع تجريبي - A"
	return [
		["1001", "موظف تجريبي", default_cost_center, "الرياض", "محاسب", "Bank", 3000, 500, 300, 200, 4000, 30, 0, 0, 0, 0, 300, 0, 200, 500, 3500, "1234567890", "GOSI-1001"],
		["1001", "موظف تجريبي", second_cost_center, "الرياض", "محاسب", "Bank", 1000, 0, 0, 0, 1000, 30, 0, 0, 0, 0, 100, 0, 0, 100, 900, "1234567890", "GOSI-1001"],
		["1002", "موظف تجريبي 2", default_cost_center, "جدة", "سائق", "Cash", 2500, 300, 200, 0, 3000, 30, 1, 100, 0, 0, 225, 150, 125, 350, 2800, "2234567890", "GOSI-1002"],
	]


def _get_second_postable_cost_center(company: str, excluded_cost_center: str) -> str:
	rows = frappe.get_all(
		"Cost Center",
		filters={"company": company, "is_group": 0},
		fields=["name"],
		order_by="lft asc",
	)
	for row in rows:
		if row["name"] != excluded_cost_center:
			return row["name"]
	return ""


def _merge_journal_entry_accounts(accounts: list[dict]) -> list[dict]:
	merged = {}
	ordered_keys = []

	for row in accounts:
		key = (
			cstr(row.get("account") or "").strip(),
			cstr(row.get("cost_center") or "").strip(),
			cstr(row.get("party_type") or "").strip(),
			cstr(row.get("party") or "").strip(),
		)
		if key not in merged:
			merged[key] = {
				"account": row.get("account"),
				"cost_center": row.get("cost_center"),
				"party_type": row.get("party_type"),
				"party": row.get("party"),
				"debit_in_account_currency": 0.0,
				"credit_in_account_currency": 0.0,
			}
			ordered_keys.append(key)

		merged[key]["debit_in_account_currency"] += flt(row.get("debit_in_account_currency"))
		merged[key]["credit_in_account_currency"] += flt(row.get("credit_in_account_currency"))

	return [merged[key] for key in ordered_keys]


# ─── Private Helpers ────────────────────────────────────────────────────────────

def _build_employee_row(emp: dict, month: str, year: int) -> dict:
	"""بناء بيانات صف الموظف الواحد في الجدول الفرعي."""
	salary = get_employee_salary_components(emp["name"])
	basic = salary["basic_salary"]
	assert_positive_basic_salary(emp.get("employee_name") or emp["name"], basic, _("building payroll row / تجهيز صف الراتب"))
	housing = salary["housing_allowance"]
	transport = salary["transport_allowance"]
	other = salary["other_allowances"]

	gross = round(basic + housing + transport + other, 2)

	# اقتطاع GOSI للموظف
	gosi_rate = flt(get_gosi_rates(emp.get("nationality") or "").get("employee_rate"))
	gosi_base = min(basic, GOSI_MAX_BASE)
	gosi_deduction = round(gosi_base * gosi_rate / 100, 2)

	# خصم الإجازة المرضية (إن وجدت في الشهر الحالي)
	month_num = _month_name_to_num(month)
	month_start = f"{year}-{month_num:02d}-01"
	import calendar
	last_day = calendar.monthrange(int(year), month_num)[1]
	month_end = f"{year}-{month_num:02d}-{last_day:02d}"

	sick_rows = frappe.get_all(
		"Saudi Sick Leave",
		filters={
			"employee": emp["name"],
			"docstatus": 1,
		},
		fields=["leave_pay_amount", "daily_salary", "total_days", "pay_rate", "from_date", "to_date"],
	)
	month_filtered_sick_rows = [
		sr for sr in sick_rows
		if getdate(sr.from_date) <= getdate(month_end) and getdate(sr.to_date) >= getdate(month_start)
	]
	sick_deduction = calculate_prorated_sick_leave_deduction(month_filtered_sick_rows, month_start, month_end, round(basic / 30, 2))

	# إضافة العمل الإضافي المعتمد في الشهر
	ot_rows = frappe.get_all(
		"Overtime Request",
		filters={
			"employee": emp["name"],
			"docstatus": 1,
			"date": ["between", [month_start, month_end]],
		},
		fields=["overtime_amount", "approval_status"],
	)
	overtime = round(sum(flt(r.overtime_amount) for r in ot_rows if text_matches_tokens(r.approval_status, "approved", "موافق")), 2)
	loan_info = get_due_loan_deduction(emp["name"], month_num, int(year))
	loan_deduction = round(flt(loan_info.get("loan_deduction")), 2)
	total_deductions = round(gosi_deduction + sick_deduction + loan_deduction, 2)

	net = round(gross + overtime - total_deductions, 2)

	return {
		"employee": emp["name"],
		"employee_name": emp.get("employee_name", ""),
		"department": emp.get("department", ""),
		"cost_center": _get_company_default_cost_center(emp.get("company") or ""),
		"nationality": emp.get("nationality", ""),
		"work_location": "",
		"designation": "",
		"salary_mode": "",
		"gosi_registration": "",
		"working_days": 0.0,
		"absence_days": 0.0,
		"late_hours": 0.0,
		"basic_salary": basic,
		"housing_allowance": housing,
		"transport_allowance": transport,
		"other_allowances": other,
		"gross_salary": gross,
		"gosi_employee_deduction": gosi_deduction,
		"sick_leave_deduction": round(sick_deduction, 2),
		"loan_deduction": loan_deduction,
		"other_deductions": 0.0,
		"total_deductions": total_deductions,
		"overtime_addition": overtime,
		"loan_installments": ", ".join(loan_info.get("installment_names", [])),
		"net_salary": net,
	}


def _extract_source_workbook_rows(content: bytes) -> list[dict]:
	workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
	worksheet = _get_source_workbook_sheet(workbook)
	header_row_index, header_map = _find_source_header_row(worksheet)
	rows = []

	for row_index, values in enumerate(
		worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
		start=header_row_index + 1,
	):
		payload = {
			fieldname: values[column_index] if column_index < len(values) else None
			for fieldname, column_index in header_map.items()
		}
		if _is_empty_source_row(payload):
			continue
		payload["source_row"] = row_index
		rows.append(payload)

	return rows


def _preview_payroll_workbook(company: str, workbook_url: str) -> dict:
	raw_rows = _extract_source_workbook_rows(_get_attached_file_content(workbook_url))
	import_rows, warnings = _map_workbook_rows_to_payroll(company, raw_rows)
	unmatched_rows = [warning for warning in warnings if "could not match employee" in warning]
	matched_fallbacks = [warning for warning in warnings if "matched employee" in warning]
	unmatched_details = _collect_unmatched_workbook_rows(company, raw_rows)
	company_employee_count = frappe.db.count("Employee", {"company": company})

	return {
		"total_rows": len(raw_rows),
		"importable_rows": len(import_rows),
		"unmatched_rows": len(unmatched_rows),
		"company_employee_count": company_employee_count,
		"sample_unmatched": unmatched_rows[:10],
		"sample_matched": matched_fallbacks[:10],
		"unmatched_details": unmatched_details,
		"warnings": warnings,
		"import_rows": import_rows,
	}


def _validate_payroll_workbook_rows(company: str, raw_rows: list[dict]) -> dict:
	lookup = _get_company_employee_lookup(company)
	errors = []
	warnings = []
	would_create_cost_centers = []
	duplicate_keys = {}
	required_fields = {
		"employee_id": _("Payroll ID / الرقم الوظيفي"),
		"employee_name": _("Name / الاسم"),
		"basic_salary": _("Basic / الأساسي"),
		"gross_salary": _("Gross / الإجمالي"),
		"total_deductions": _("Total Deductions / إجمالي الخصومات"),
		"net_salary": _("Net / الصافي"),
	}

	for raw in raw_rows:
		row_label = raw.get("source_row") or "?"
		missing_fields = []
		for fieldname, label in required_fields.items():
			if _is_blank(raw.get(fieldname)):
				missing_fields.append(label)
		if _is_blank(raw.get("cost_center")) and _is_blank(raw.get("department")):
			missing_fields.append(_("Cost Center / مركز التكلفة"))
		if missing_fields:
			errors.append(_("الصف {0}: توجد بيانات إلزامية ناقصة: {1}.").format(row_label, ", ".join(missing_fields)))

		cost_center_label = cstr(raw.get("cost_center") or raw.get("department") or "").strip()
		employee_key = cstr(raw.get("employee_id") or raw.get("employee_name") or "").strip().lower()
		duplicate_key = (employee_key, cost_center_label.lower())
		if employee_key and cost_center_label:
			if duplicate_key in duplicate_keys:
				errors.append(_("الصف {0}: الموظف مكرر على مركز التكلفة نفسه كما في الصف {1}.").format(row_label, duplicate_keys[duplicate_key]))
			else:
				duplicate_keys[duplicate_key] = row_label

		basic = _to_currency(raw.get("basic_salary"))
		gross = _to_currency(raw.get("gross_salary"))
		total_deductions = _to_currency(raw.get("total_deductions"))
		gosi = _to_currency(raw.get("gosi_deduction"))
		overtime = _to_currency(raw.get("additions"))
		manual_deduction = _to_currency(raw.get("manual_deduction"))
		absence_value = _to_currency(raw.get("absence_value"))
		late_value = _to_currency(raw.get("late_value"))
		net = _to_currency(raw.get("net_salary"))
		if basic < 0 or gross < 0 or total_deductions < 0 or net < 0:
			errors.append(_("الصف {0}: لا يمكن أن تحتوي قيم الرواتب أو الخصومات على أرقام سالبة.").format(row_label))

		calculated_deductions = total_deductions or round(gosi + manual_deduction + absence_value + late_value, 2)
		calculated_net = round(gross + overtime - calculated_deductions, 2)
		if not _is_blank(raw.get("net_salary")) and abs(net - calculated_net) > 0.01:
			errors.append(_("الصف {0}: صافي الراتب لا يطابق المعادلة المحاسبية. المدخل {1:.2f} والمتوقع {2:.2f}.").format(row_label, net, calculated_net))

		employee, matched_by = _match_workbook_employee(raw, lookup)
		if not employee:
			warnings.append(_("الصف {0}: الموظف {1} غير مرتبط حالياً بسجل Employee موجود داخل الشركة.").format(row_label, raw.get("employee_id") or raw.get("employee_name") or _("غير محدد")))
		elif matched_by != "employee_id":
			warnings.append(_("الصف {0}: تمت مطابقة الموظف باستخدام {1} بدلاً من الرقم الوظيفي.").format(row_label, matched_by))

		if cost_center_label and not _get_postable_cost_center(cost_center_label, company):
			would_create_cost_centers.append(cost_center_label)

	return {
		"total_rows": len(raw_rows),
		"error_count": len(errors),
		"warning_count": len(warnings),
		"errors": errors,
		"warnings": warnings,
		"would_create_cost_centers": sorted({value for value in would_create_cost_centers if value}),
	}


def _collect_unmatched_workbook_rows(company: str, raw_rows: list[dict]) -> list[dict]:
	lookup = _get_company_employee_lookup(company)
	missing = []
	for raw in raw_rows:
		employee, _matched_by = _match_workbook_employee(raw, lookup)
		if employee:
			continue
		missing.append(raw)
	return missing


def _build_gap_report_rows(unmatched_rows: list[dict]) -> list[list]:
	rows = [[
		"source_row",
		"employee_id",
		"employee_name",
		"national_id",
		"department",
		"work_location",
		"designation",
		"gross_salary",
		"net_salary",
		"reason",
	]]

	for row in unmatched_rows:
		rows.append([
			row.get("source_row"),
			row.get("employee_id"),
			row.get("employee_name"),
			row.get("national_id"),
			row.get("department"),
			row.get("work_location"),
			row.get("designation"),
			row.get("gross_salary"),
			row.get("net_salary"),
			"No matching Employee record in selected company",
		])

	return rows


def _build_employee_setup_template_rows(company: str, unmatched_rows: list[dict]) -> list[list]:
	rows = [EMPLOYEE_SETUP_TEMPLATE_HEADERS]

	for row in unmatched_rows:
		payroll_employee_id = cstr(row.get("employee_id") or "").strip()
		first_name, middle_name, last_name = _split_payroll_employee_name(row.get("employee_name"))
		rows.append([
			row.get("source_row"),
			payroll_employee_id,
			row.get("employee_name"),
			row.get("national_id"),
			company,
			row.get("department"),
			row.get("designation"),
			payroll_employee_id,
			first_name,
			middle_name,
			last_name,
			"",
			"",
			"",
			"Active",
			"Fill the required HR fields before import",
		])

	return rows


def _extract_employee_setup_rows(content: bytes) -> list[dict]:
	workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
	worksheet = workbook["Employee Setup"] if "Employee Setup" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
	rows = list(worksheet.iter_rows(values_only=True))
	if not rows:
		return []

	header_map = {
		_normalize_header_text(value): index
		for index, value in enumerate(rows[0])
		if _normalize_header_text(value)
	}
	missing_headers = [header for header in EMPLOYEE_SETUP_TEMPLATE_HEADERS if header not in header_map]
	if missing_headers:
		frappe.throw(
			_(
				"The employee setup workbook is missing required columns: {0}<br>"
				"ملف إعداد الموظفين ينقصه الأعمدة المطلوبة: {0}"
			).format(", ".join(missing_headers))
		)

	payloads = []
	for row_values in rows[1:]:
		payload = {
			header: row_values[column_index] if column_index < len(row_values) else None
			for header, column_index in header_map.items()
			if header in EMPLOYEE_SETUP_TEMPLATE_HEADERS
		}
		if _is_empty_employee_setup_row(payload):
			continue
		payloads.append(payload)

	return payloads


def _build_employee_setup_workbook_rows(rows: list[dict]) -> list[list]:
	output = [EMPLOYEE_SETUP_TEMPLATE_HEADERS]
	for row in rows:
		output.append([row.get(header) for header in EMPLOYEE_SETUP_TEMPLATE_HEADERS])
	return output


def _autofill_employee_setup_names(rows: list[dict]) -> tuple[list[dict], int]:
	updated = 0
	for row in rows:
		first_name, middle_name, last_name = _split_payroll_employee_name(row.get("payroll_employee_name"))
		row_updated = False
		if first_name and not cstr(row.get("first_name") or "").strip():
			row["first_name"] = first_name
			row_updated = True
		if middle_name and not cstr(row.get("middle_name") or "").strip():
			row["middle_name"] = middle_name
			row_updated = True
		if last_name and not cstr(row.get("last_name") or "").strip():
			row["last_name"] = last_name
			row_updated = True
		if row_updated:
			updated += 1
	return rows, updated


def _split_payroll_employee_name(value) -> tuple[str, str, str]:
	name = " ".join(cstr(value or "").replace("\n", " ").replace("\t", " ").split()).strip()
	if not name:
		return "", "", ""

	parts = [part for part in name.split(" ") if part]
	if len(parts) == 1:
		return parts[0], "", ""
	if len(parts) == 2:
		return parts[0], "", parts[1]
	return parts[0], " ".join(parts[1:-1]), parts[-1]


def _is_empty_employee_setup_row(payload: dict) -> bool:
	key_fields = [
		payload.get("employee_number") or payload.get("payroll_employee_id"),
		payload.get("payroll_employee_name"),
		payload.get("first_name"),
	]
	return all(_is_blank(value) for value in key_fields)


def _import_employee_setup_rows(company: str, rows: list[dict]) -> tuple[list[str], list[str]]:
	if not rows:
		frappe.throw(_("The employee setup workbook does not contain any rows.<br>ملف إعداد الموظفين لا يحتوي على أي صفوف."))

	prepared_rows = []
	skipped = []
	for row in rows:
		prepared = _prepare_employee_setup_row(company, row)
		if not prepared:
			skipped.append(
				_(
					"Row {0}: employee {1} already exists and was skipped."
				).format(row.get("source_row") or "?", row.get("employee_number") or row.get("payroll_employee_id") or row.get("payroll_employee_name") or "N/A")
			)
			continue
		prepared_rows.append(prepared)

	created = []
	savepoint = "saudi_hr_employee_setup_import"
	frappe.db.savepoint(savepoint)
	try:
		for payload in prepared_rows:
			employee = frappe.get_doc(payload)
			employee.insert()
			created.append(employee.name)
	except Exception:
		frappe.db.rollback(save_point=savepoint)
		raise
	else:
		frappe.db.release_savepoint(savepoint)

	return created, skipped


def _prepare_employee_setup_row(company: str, row: dict) -> dict | None:
	meta = frappe.get_meta("Employee")
	employee_number = cstr(row.get("employee_number") or row.get("payroll_employee_id") or "").strip()
	row_label = row.get("source_row") or "?"
	if not employee_number:
		frappe.throw(_("Row {0}: employee_number is required.<br>الصف {0}: الرقم الوظيفي مطلوب.").format(row_label))

	if frappe.db.exists("Employee", {"company": company, "employee_number": employee_number}):
		return None

	name_parts = _split_payroll_employee_name(row.get("payroll_employee_name"))
	first_name = cstr(row.get("first_name") or name_parts[0] or "").strip()
	middle_name = cstr(row.get("middle_name") or name_parts[1] or "").strip()
	last_name = cstr(row.get("last_name") or name_parts[2] or "").strip()
	gender = cstr(row.get("gender") or "").strip()
	status = cstr(row.get("status") or "Active").strip()
	if not first_name:
		frappe.throw(_("Row {0}: first_name is required.<br>الصف {0}: الاسم الأول مطلوب.").format(row_label))
	if not gender:
		frappe.throw(_("Row {0}: gender is required.<br>الصف {0}: الجنس مطلوب.").format(row_label))
	if frappe.db.exists("DocType", "Gender") and not frappe.db.exists("Gender", gender):
		frappe.throw(_("Row {0}: gender was not found.<br>الصف {0}: الجنس غير موجود.").format(row_label))

	try:
		date_of_birth = getdate(row.get("date_of_birth"))
	except Exception:
		frappe.throw(_("Row {0}: date_of_birth is invalid.<br>الصف {0}: تاريخ الميلاد غير صالح.").format(row_label))

	try:
		date_of_joining = getdate(row.get("date_of_joining"))
	except Exception:
		frappe.throw(_("Row {0}: date_of_joining is invalid.<br>الصف {0}: تاريخ المباشرة غير صالح.").format(row_label))

	today_date = getdate(today())
	if date_of_birth > today_date:
		frappe.throw(_("Row {0}: date_of_birth cannot be in the future.<br>الصف {0}: تاريخ الميلاد لا يمكن أن يكون في المستقبل.").format(row_label))
	if date_of_joining > today_date:
		frappe.throw(_("Row {0}: date_of_joining cannot be in the future.<br>الصف {0}: تاريخ المباشرة لا يمكن أن يكون في المستقبل.").format(row_label))
	if date_of_joining <= date_of_birth:
		frappe.throw(_("Row {0}: date_of_joining must be after date_of_birth.<br>الصف {0}: تاريخ المباشرة يجب أن يكون بعد تاريخ الميلاد.").format(row_label))

	payload = {
		"doctype": "Employee",
		"company": cstr(row.get("company") or company).strip() or company,
		"employee_number": employee_number,
		"first_name": first_name,
		"gender": gender,
		"date_of_birth": date_of_birth,
		"date_of_joining": date_of_joining,
		"status": status,
	}

	if payload["company"] != company:
		frappe.throw(
			_("Row {0}: company must match the payroll company {1}.<br>الصف {0}: يجب أن تطابق الشركة شركة كشف الرواتب {1}.").format(row_label, company)
		)

	for fieldname, value in (
		("middle_name", middle_name),
		("last_name", last_name),
		("department", cstr(row.get("department") or "").strip()),
		("designation", cstr(row.get("designation") or "").strip()),
	):
		if fieldname == "department" and value and not frappe.db.exists("Department", value):
			frappe.throw(_("Row {0}: department was not found.<br>الصف {0}: القسم غير موجود.").format(row_label))
		if fieldname == "designation" and value and not frappe.db.exists("Designation", value):
			frappe.throw(_("Row {0}: designation was not found.<br>الصف {0}: المسمى الوظيفي غير موجود.").format(row_label))
		if value and meta.has_field(fieldname):
			payload[fieldname] = value

	return payload


def _normalize_basic_employee_creation_defaults(
	default_gender: str,
	default_date_of_birth,
	default_date_of_joining,
	default_status: str,
) -> dict:
	gender = cstr(default_gender or "").strip()
	if not gender:
		frappe.throw(_("Default gender is required.<br>الجنس الافتراضي مطلوب."))
	if frappe.db.exists("DocType", "Gender") and not frappe.db.exists("Gender", gender):
		frappe.throw(_("Default gender was not found.<br>الجنس الافتراضي غير موجود."))

	try:
		date_of_birth = getdate(default_date_of_birth)
	except Exception:
		frappe.throw(_("Default date of birth is invalid.<br>تاريخ الميلاد الافتراضي غير صالح."))

	try:
		date_of_joining = getdate(default_date_of_joining)
	except Exception:
		frappe.throw(_("Default date of joining is invalid.<br>تاريخ المباشرة الافتراضي غير صالح."))

	today_date = getdate(today())
	if date_of_birth > today_date:
		frappe.throw(_("Default date of birth cannot be in the future.<br>تاريخ الميلاد الافتراضي لا يمكن أن يكون في المستقبل."))
	if date_of_joining > today_date:
		frappe.throw(_("Default date of joining cannot be in the future.<br>تاريخ المباشرة الافتراضي لا يمكن أن يكون في المستقبل."))
	if date_of_joining <= date_of_birth:
		frappe.throw(_("Default date of joining must be after date of birth.<br>تاريخ المباشرة الافتراضي يجب أن يكون بعد تاريخ الميلاد."))

	status = cstr(default_status or "Active").strip() or "Active"
	return {
		"gender": gender,
		"date_of_birth": date_of_birth,
		"date_of_joining": date_of_joining,
		"status": status,
	}


def _create_basic_employees_for_payroll(doc, defaults: dict) -> tuple[list[str], int, list[str]]:
	lookup = _get_company_employee_lookup(doc.company)
	created = []
	linked = 0
	skipped = []
	savepoint = "saudi_hr_basic_employee_create"
	frappe.db.savepoint(savepoint)

	try:
		for row in doc.employees:
			if cstr(row.employee or "").strip():
				continue

			raw = {
				"employee_id": getattr(row, "payroll_employee_id", None),
				"employee_name": getattr(row, "employee_name", None),
			}
			has_payroll_employee_id = bool(cstr(getattr(row, "payroll_employee_id", "") or "").strip())
			employee, _matched_by = _match_workbook_employee(
				raw,
				lookup,
				allow_related_employee_id=not has_payroll_employee_id,
				allow_name_match=not has_payroll_employee_id,
			)
			if employee:
				_hydrate_payroll_row_from_employee(row, employee)
				linked += 1
				continue

			if not cstr(getattr(row, "payroll_employee_id", "") or "").strip() and not cstr(getattr(row, "employee_name", "") or "").strip():
				skipped.append(
					_("Row {0}: missing payroll employee id and employee name.").format(row.idx or "?")
				)
				continue

			payload = _build_basic_employee_payload_from_payroll_row(doc.company, row, defaults)
			employee_doc = frappe.get_doc(payload)
			employee_doc.insert()

			employee_data = {
				"name": employee_doc.name,
				"employee_name": employee_doc.get("employee_name") or cstr(getattr(row, "employee_name", "") or "").strip(),
				"department": employee_doc.get("department") or "",
				"nationality": employee_doc.get("nationality") or "",
				"employee_number": employee_doc.get("employee_number") or payload.get("employee_number") or "",
			}

			for source, matched_by in (
				(employee_data.get("name"), "name"),
				(employee_data.get("employee_number"), "employee_id"),
				(employee_data.get("employee_name"), "employee_name"),
			):
				_merge_employee_lookup(lookup, employee_data, source, matched_by)

			_hydrate_payroll_row_from_employee(row, employee_data)
			created.append(employee_doc.name)
	except Exception:
		frappe.db.rollback(save_point=savepoint)
		raise
	else:
		frappe.db.release_savepoint(savepoint)

	return created, linked, skipped


def _build_basic_employee_payload_from_payroll_row(company: str, row, defaults: dict) -> dict:
	meta = frappe.get_meta("Employee")
	employee_number = cstr(getattr(row, "payroll_employee_id", "") or "").strip()
	first_name, middle_name, last_name = _split_payroll_employee_name(getattr(row, "employee_name", None))
	first_name = first_name or employee_number or _("Payroll")

	payload = {
		"doctype": "Employee",
		"company": company,
		"first_name": first_name,
		"gender": defaults["gender"],
		"date_of_birth": defaults["date_of_birth"],
		"date_of_joining": defaults["date_of_joining"],
		"status": defaults["status"],
	}

	if employee_number:
		payload["employee_number"] = employee_number

	for fieldname, value in (
		("middle_name", middle_name),
		("last_name", last_name),
		("department", _resolve_department_link(getattr(row, "department", None))),
		("designation", _resolve_designation_link(getattr(row, "designation", None))),
	):
		if value and meta.has_field(fieldname):
			payload[fieldname] = value

	return payload


def _hydrate_payroll_row_from_employee(row, employee: dict):
	row.employee = employee.get("name") or row.employee
	row.employee_name = employee.get("employee_name") or row.employee_name
	department = _resolve_department_link(employee.get("department") or getattr(row, "department", None))
	if department:
		row.department = department
	row.cost_center = _resolve_cost_center_link(getattr(row, "cost_center", None), employee.get("company") or "") or getattr(row, "cost_center", None)
	if employee.get("nationality"):
		row.nationality = employee.get("nationality")


def _get_source_workbook_sheet(workbook):
	for sheet_name in PREFERRED_SOURCE_WORKBOOK_SHEETS:
		if sheet_name not in workbook.sheetnames:
			continue

		worksheet = workbook[sheet_name]
		try:
			_find_source_header_row(worksheet)
			return worksheet
		except frappe.ValidationError:
			continue

	for worksheet in workbook.worksheets:
		try:
			_find_source_header_row(worksheet)
			return worksheet
		except frappe.ValidationError:
			continue

	frappe.throw(
		_("Could not locate the source payroll sheet in the workbook.<br>تعذّر العثور على شيت الرواتب المصدر داخل الملف."),
		title=_("Sheet Not Found / الشيت غير موجود"),
	)


def _find_source_header_row(worksheet):
	for row_index, values in enumerate(worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
		header_map = {}
		for column_index, cell_value in enumerate(values):
			canonical = _canonical_workbook_header(cell_value)
			if canonical and canonical not in header_map:
				header_map[canonical] = column_index

		if {"employee_id", "employee_name", "basic_salary", "net_salary"}.issubset(header_map):
			return row_index, header_map

	frappe.throw(
		_("Could not find the payroll header row in the workbook.<br>تعذّر العثور على صف رؤوس الرواتب داخل الملف."),
		title=_("Header Not Found / لم يتم العثور على الرؤوس"),
	)


def _canonical_workbook_header(value) -> str | None:
	normalized = _normalize_header_text(value)
	if not normalized:
		return None

	for fieldname, aliases in WORKBOOK_HEADER_ALIASES.items():
		if normalized in {_normalize_header_text(alias) for alias in aliases}:
			return fieldname

	return None


def _normalize_header_text(value) -> str:
	return " ".join(cstr(value or "").replace("\n", " ").replace("\r", " ").split()).strip()


def _is_empty_source_row(payload: dict) -> bool:
	if _is_blank(payload.get("employee_id")) and _is_blank(payload.get("employee_name")):
		return True

	key_fields = [
		payload.get("employee_id"),
		payload.get("employee_name"),
		payload.get("gross_salary"),
		payload.get("net_salary"),
	]
	return all(_is_blank(value) for value in key_fields)


def _map_workbook_rows_to_payroll(company: str, raw_rows: list[dict]) -> tuple[list[dict], list[str]]:
	employees_by_key = _get_company_employee_lookup(company)
	import_rows = []
	warnings = []

	for raw in raw_rows:
		employee, matched_by = _match_workbook_employee(raw, employees_by_key)
		if not employee:
			warnings.append(
				_(f"Row {raw.get('source_row')}: could not match employee {raw.get('employee_id') or raw.get('employee_name')}.")
			)

		basic = _to_currency(raw.get("basic_salary"))
		housing = _to_currency(raw.get("housing_allowance"))
		transport = _to_currency(raw.get("transport_allowance"))
		other_allowances = _to_currency(raw.get("other_allowances"))
		component_gross = round(basic + housing + transport + other_allowances, 2)
		gosi = _to_currency(raw.get("gosi_deduction"))
		overtime = _to_currency(raw.get("additions"))
		gross = _normalize_workbook_gross_salary(
			_to_currency(raw.get("gross_salary")),
			component_gross,
			overtime,
		)
		total_deductions = _to_currency(raw.get("total_deductions"))
		if not total_deductions:
			total_deductions = round(
				gosi + _to_currency(raw.get("manual_deduction")) + _to_currency(raw.get("absence_value")) + _to_currency(raw.get("late_value")),
				2,
			)

		other_deductions = round(max(total_deductions - gosi, 0.0), 2)
		net = _to_currency(raw.get("net_salary"))
		if not net:
			net = round(gross + overtime - gosi - other_deductions, 2)

		calculated_net = round(gross + overtime - total_deductions, 2)
		if net and abs(net - calculated_net) > 0.01:
			warnings.append(
				_(
					f"Row {raw.get('source_row')}: net salary mismatch in workbook for "
					f"{(employee or {}).get('name') or raw.get('employee_id') or raw.get('employee_name')}; "
					f"imported net {net:.2f} and calculated net {calculated_net:.2f}."
				)
			)

		import_rows.append({
			"payroll_employee_id": cstr(raw.get("employee_id") or "").strip(),
			"employee": (employee or {}).get("name") or "",
			"employee_name": (employee or {}).get("employee_name") or raw.get("employee_name") or "",
			"workbook_department": cstr(raw.get("department") or raw.get("cost_center") or "").strip(),
			"work_location": cstr(raw.get("work_location") or "").strip(),
			"designation": cstr(raw.get("designation") or "").strip(),
			"salary_mode": cstr(raw.get("salary_mode") or "").strip(),
			"gosi_registration": cstr(raw.get("gosi_registration") or "").strip(),
			"department": _resolve_department_link((employee or {}).get("department")),
			"cost_center": _resolve_cost_center_link(raw.get("cost_center") or raw.get("department"), company),
			"nationality": (employee or {}).get("nationality") or "",
			"working_days": _to_number(raw.get("working_days")),
			"absence_days": _to_number(raw.get("absence_days")),
			"late_hours": _to_number(raw.get("late_hours")),
			"basic_salary": basic,
			"housing_allowance": housing,
			"transport_allowance": transport,
			"other_allowances": other_allowances,
			"gross_salary": gross,
			"gosi_employee_deduction": gosi,
			"sick_leave_deduction": 0.0,
			"loan_deduction": 0.0,
			"other_deductions": other_deductions,
			"total_deductions": round(gosi + other_deductions, 2),
			"overtime_addition": overtime,
			"loan_installments": "",
			"net_salary": net,
		})

		if not employee:
			warnings.append(
				_(
					f"Row {raw.get('source_row')}: imported from workbook without a linked Employee record. "
					"Use Employee Setup only if you need permanent HR records for this payroll row."
				)
			)
		elif matched_by != "employee_id":
			warnings.append(
				_(f"Row {raw.get('source_row')}: matched employee {employee['name']} by {matched_by}.")
			)

	return import_rows, warnings


def _resolve_department_link(value) -> str:
	department = cstr(value or "").strip()
	if not department:
		return ""
	return department if frappe.db.exists("Department", department) else ""


def _resolve_designation_link(value) -> str:
	designation = cstr(value or "").strip()
	if not designation:
		return ""
	return designation if frappe.db.exists("Designation", designation) else ""


def _normalize_workbook_gross_salary(raw_gross: float, component_gross: float, overtime: float) -> float:
	if raw_gross:
		if overtime and abs(raw_gross - round(component_gross + overtime, 2)) <= 0.01:
			return component_gross
		return raw_gross
	return component_gross


def _get_company_employee_lookup(company: str) -> dict[str, dict]:
	fields = ["name", "employee_name", "department", "employee_number", "company"]
	meta = frappe.get_meta("Employee")
	for optional_field in ("nationality", "passport_number", "iqama_number"):
		if meta.has_field(optional_field):
			fields.append(optional_field)

	employees = frappe.get_all("Employee", filters={"company": company}, fields=fields)
	lookup = {}

	for employee in employees:
		for source, matched_by in (
			(employee.get("name"), "name"),
			(employee.get("employee_number"), "employee_id"),
			(employee.get("passport_number"), "passport_number"),
			(employee.get("iqama_number"), "iqama_number"),
			(employee.get("employee_name"), "employee_name"),
		):
			_merge_employee_lookup(lookup, employee, source, matched_by)

	contract_rows = frappe.get_all(
		"Saudi Employment Contract",
		filters={"company": company},
		fields=["employee", "iqama_number", "passport_number"],
	)
	for row in contract_rows:
		employee = next((item for item in employees if item["name"] == row.get("employee")), None)
		if not employee:
			continue
		_merge_employee_lookup(lookup, employee, row.get("iqama_number"), "contract_iqama_number")
		_merge_employee_lookup(lookup, employee, row.get("passport_number"), "contract_passport_number")

	if frappe.db.exists("DocType", "Work Permit Iqama"):
		permit_rows = frappe.get_all(
			"Work Permit Iqama",
			filters={"company": company},
			fields=["employee", "iqama_number"],
		)
		for row in permit_rows:
			employee = next((item for item in employees if item["name"] == row.get("employee")), None)
			if not employee:
				continue
			_merge_employee_lookup(lookup, employee, row.get("iqama_number"), "work_permit_iqama_number")

	return lookup


def _merge_employee_lookup(lookup: dict[str, dict], employee: dict, source, matched_by: str):
	key = _normalize_lookup_key(source)
	if not key:
		return
	existing = lookup.get(key)
	if existing and existing["employee"] and existing["employee"]["name"] != employee["name"]:
		lookup[key] = {"employee": None, "matched_by": matched_by}
		return
	lookup[key] = {"employee": employee, "matched_by": matched_by}


def _match_workbook_employee(
	raw: dict,
	lookup: dict[str, dict],
	*,
	allow_related_employee_id: bool = True,
	allow_name_match: bool = True,
):
	for candidate, matched_by, allow_related_key in (
		(raw.get("employee_id"), "employee_id", allow_related_employee_id),
		(raw.get("national_id"), "iqama_number", False),
		(raw.get("employee_name"), "employee_name", False),
	):
		if matched_by == "employee_name" and not allow_name_match:
			continue

		for key in _candidate_lookup_keys(candidate, allow_related_key=allow_related_key):
			match = lookup.get(key)
			if match and match.get("employee"):
				return match["employee"], match.get("matched_by") or matched_by

	return None, None


def _candidate_lookup_keys(value, allow_related_key: bool = True) -> list[str]:
	key = _normalize_lookup_key(value)
	if not key:
		return []

	candidates = [key]
	base_key = key.rstrip("abcdefghijklmnopqrstuvwxyz")
	if allow_related_key and base_key and base_key != key and any(char.isdigit() for char in base_key):
		candidates.append(base_key)

	return candidates


def _normalize_lookup_key(value) -> str:
	if _is_blank(value):
		return ""
	text = cstr(value).strip().replace("\n", " ")
	if text.lower() == "blank":
		return ""
	try:
		number = float(text)
		if number.is_integer():
			return str(int(number))
	except Exception:
		pass
	return " ".join(text.split()).lower()


def _to_currency(value) -> float:
	return round(_to_number(value), 2)


def _to_number(value) -> float:
	if _is_blank(value):
		return 0.0
	if isinstance(value, str) and value.strip().lower() == "blank":
		return 0.0
	return flt(value)


def _is_blank(value) -> bool:
	return value in (None, "")


def _get_attached_file_content(file_url: str) -> bytes:
	file_row = frappe.db.get_value("File", {"file_url": file_url}, ["name", "file_name", "file_size"], as_dict=True)
	if not file_row:
		frappe.throw(_("Unable to find the uploaded payroll workbook.<br>تعذّر العثور على ملف الرواتب المرفوع."))
	file_extension = splitext(cstr(file_row.file_name or file_url).strip())[1].lower()
	if file_extension not in ALLOWED_WORKBOOK_EXTENSIONS:
		frappe.throw(
			_("Only Excel workbook files are supported.<br>الملفات المدعومة هي ملفات Excel فقط."),
			title=_("Invalid File Type / نوع ملف غير صالح"),
		)
	if flt(file_row.file_size) > MAX_WORKBOOK_FILE_SIZE_BYTES:
		frappe.throw(
			_("The uploaded workbook is too large. Please keep it under 10 MB.<br>ملف الرواتب كبير جداً. يرجى أن يكون أقل من 10 ميجابايت."),
			title=_("Workbook Too Large / ملف كبير جداً"),
		)
	return frappe.get_doc("File", file_row.name).get_content()


def _add_payroll_audit_comment(doc, message: str):
	if not cstr(message or "").strip():
		return
	doc.add_comment("Comment", text=message)


def _month_name_to_num(month_label: str) -> int:
	"""تحويل اسم الشهر الثنائي إلى رقم."""
	MONTHS = {
		"january": 1, "february": 2, "march": 3, "april": 4,
		"may": 5, "june": 6, "july": 7, "august": 8,
		"september": 9, "october": 10, "november": 11, "december": 12,
		"يناير": 1, "فبراير": 2, "مارس": 3, "أبريل": 4,
		"مايو": 5, "يونيو": 6, "يوليو": 7, "أغسطس": 8,
		"سبتمبر": 9, "أكتوبر": 10, "نوفمبر": 11, "ديسمبر": 12,
	}
	# الاسم قد يكون "January / يناير"
	for part in (month_label or "").replace("/", " ").split():
		key = part.strip().lower()
		if key in MONTHS:
			return MONTHS[key]
	return 1


def _get_employee_fetch_fields() -> list[str]:
	fields = ["name", "employee_name", "department", "company"]
	if frappe.get_meta("Employee").has_field("nationality"):
		fields.append("nationality")
	return fields


def _get_payable_account(company: str) -> str:
	"""الحصول على حساب الرواتب المستحقة للشركة."""
	account = frappe.db.get_value(
		"Account",
		{"company": company, "account_type": "Payable", "is_group": 0},
		"name",
	)
	return account or ""
