from io import BytesIO
from unittest.mock import patch
from types import SimpleNamespace

import frappe
from frappe.tests.utils import FrappeTestCase
from frappe.utils import getdate
from openpyxl import Workbook, load_workbook

from saudi_hr.saudi_hr.doctype.saudi_monthly_payroll import saudi_monthly_payroll as payroll_module


class TestSaudiMonthlyPayroll(FrappeTestCase):
	def test_build_employee_row_includes_loan_in_total_deductions(self):
		emp = {"name": "EMP-0001", "employee_name": "Demo Employee", "department": "HR", "nationality": "Saudi"}
		with patch.object(
			payroll_module, "get_employee_salary_components", return_value={
				"basic_salary": 1000,
				"housing_allowance": 200,
				"transport_allowance": 100,
				"other_allowances": 50,
			}
		), patch.object(payroll_module.frappe, "get_all", side_effect=[[], []]), patch.object(
			payroll_module, "get_gosi_rates", return_value={"employee_rate": 10.0}
		), patch.object(
			payroll_module, "get_due_loan_deduction", return_value={"loan_deduction": 125, "installment_names": ["INST-1"]}
		):
			row = payroll_module._build_employee_row(emp, "March / مارس", 2026)

		self.assertEqual(row["gross_salary"], 1350)
		self.assertEqual(row["gosi_employee_deduction"], 100)
		self.assertEqual(row["loan_deduction"], 125)
		self.assertEqual(row["total_deductions"], 225)
		self.assertEqual(row["net_salary"], 1125)
		self.assertEqual(row["loan_installments"], "INST-1")

	def test_build_employee_row_rejects_zero_basic_salary(self):
		emp = {"name": "EMP-0001", "employee_name": "Demo Employee", "department": "HR", "nationality": "Saudi"}
		with patch.object(
			payroll_module, "get_employee_salary_components", return_value={
				"basic_salary": 0,
				"housing_allowance": 0,
				"transport_allowance": 0,
				"other_allowances": 0,
			}
		):
			with self.assertRaises(frappe.ValidationError):
				payroll_module._build_employee_row(emp, "March / مارس", 2026)

	def test_build_employee_row_prorates_sick_leave_overlap(self):
		emp = {"name": "EMP-0001", "employee_name": "Demo Employee", "department": "HR", "nationality": "Saudi"}
		with patch.object(
			payroll_module, "get_employee_salary_components", return_value={
				"basic_salary": 3000,
				"housing_allowance": 0,
				"transport_allowance": 0,
				"other_allowances": 0,
			}
		), patch.object(payroll_module, "get_gosi_rates", return_value={"employee_rate": 0.0}), patch.object(
			payroll_module.frappe, "get_all", side_effect=[
				[
					frappe._dict({
						"leave_pay_amount": 350,
						"daily_salary": 100,
						"total_days": 7,
						"from_date": "2026-02-27",
						"to_date": "2026-03-05",
					})
				],
				[],
			]
		), patch.object(payroll_module, "get_due_loan_deduction", return_value={"loan_deduction": 0, "installment_names": []}):
			row = payroll_module._build_employee_row(emp, "March / مارس", 2026)

		self.assertEqual(row["sick_leave_deduction"], 250)
		self.assertEqual(row["net_salary"], 2750)

	def test_fetch_employees_skips_zero_basic_salary_and_returns_warning(self):
		class FakePayrollDoc:
			def __init__(self):
				self.company = "amd"
				self.month = "March / مارس"
				self.year = 2026
				self.employees = []
				self.total_net_payable = 0.0
				self.saved = False
				self.comments = []

			def set(self, fieldname, value):
				setattr(self, fieldname, value)

			def append(self, fieldname, value):
				getattr(self, fieldname).append(value)

			def _recalculate_totals(self):
				self.total_net_payable = round(sum(row.get("net_salary", 0.0) for row in self.employees), 2)

			def save(self):
				self.saved = True

			def add_comment(self, comment_type, text):
				self.comments.append((comment_type, text))

		fake_doc = FakePayrollDoc()
		employees = [
			{"name": "EMP-VALID", "employee_name": "Valid Employee", "department": "HR", "company": "amd", "nationality": "Saudi"},
			{"name": "EMP-ZERO", "employee_name": "Zero Employee", "department": "HR", "company": "amd", "nationality": "Saudi"},
		]
		contract_rows = [
			{
				"employee": "EMP-VALID",
				"basic_salary": 5000,
				"housing_allowance": 500,
				"transport_allowance": 250,
				"other_allowances": 100,
				"total_salary": 5850,
			},
			{
				"employee": "EMP-ZERO",
				"basic_salary": 0,
				"housing_allowance": 0,
				"transport_allowance": 0,
				"other_allowances": 0,
				"total_salary": 0,
			},
		]

		with patch.object(payroll_module.frappe, "get_doc", return_value=fake_doc), \
		     patch.object(payroll_module.frappe, "has_permission"), \
		     patch.object(payroll_module.frappe, "get_all", return_value=employees), \
		     patch.object(payroll_module.frappe.db, "sql", side_effect=[contract_rows, [], []]), \
		     patch.object(payroll_module, "_get_employee_fetch_fields", return_value=["name", "employee_name", "department", "company", "nationality"]), \
		     patch.object(payroll_module, "get_employee_salary_components", side_effect=[{"basic_salary": 0}]), \
		     patch.object(payroll_module, "get_gosi_rates", return_value={"employee_rate": 10.0}), \
		     patch.object(payroll_module, "get_due_loan_deduction", return_value={"loan_deduction": 0, "installment_names": []}), \
		     patch.object(payroll_module, "_get_company_default_cost_center", return_value="Main - AMD"):
			result = payroll_module.fetch_employees("SAU-PAY-TEST")

		self.assertTrue(fake_doc.saved)
		self.assertEqual(result["count"], 1)
		self.assertEqual(result["source_count"], 2)
		self.assertEqual(result["skipped_count"], 1)
		self.assertEqual(len(fake_doc.employees), 1)
		self.assertEqual(fake_doc.employees[0]["employee"], "EMP-VALID")
		self.assertEqual(fake_doc.total_net_payable, 5350.0)
		self.assertIn("Zero Employee", result["warnings"][0])
		self.assertIn("skipped 1", fake_doc.comments[0][1])

	def test_recalculate_totals_includes_loan_and_sick_deductions(self):
		doc = frappe.get_doc({
			"doctype": "Saudi Monthly Payroll",
			"company": "amd",
			"month": "March / مارس",
			"year": 2026,
			"posting_date": "2026-03-31",
			"employees": [
				{"employee": "EMP-1", "gross_salary": 1000, "gosi_employee_deduction": 100, "sick_leave_deduction": 20, "loan_deduction": 50, "other_deductions": 30, "overtime_addition": 10, "net_salary": 810},
				{"employee": "EMP-2", "gross_salary": 1200, "gosi_employee_deduction": 120, "sick_leave_deduction": 0, "loan_deduction": 80, "other_deductions": 25, "overtime_addition": 0, "net_salary": 975},
			],
		})

		doc._recalculate_totals()

		self.assertEqual(doc.total_employees, 2)
		self.assertEqual(doc.total_gross, 2200)
		self.assertEqual(doc.total_gosi_deductions, 220)
		self.assertEqual(doc.total_sick_deductions, 20)
		self.assertEqual(doc.total_loan_deductions, 130)
		self.assertEqual(doc.total_other_deductions, 55)
		self.assertEqual(doc.total_overtime, 10)
		self.assertEqual(doc.total_net_payable, 1785)

	def test_extract_source_workbook_rows_maps_expected_headers(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = "كشف المصدر"
		worksheet.append([None])
		worksheet.append([None])
		worksheet.append([None])
		worksheet.append(["شركة اختبار"])
		worksheet.append([
			"التسلسل",
			"الرقم الوظيفي",
			"الاسم",
			"الوظيفة",
			"مكان العمل",
			"الإدارة",
			"بنك كاش",
			"الاساسي",
			"بدل السكن",
			"بدل المواصلات",
			"بدلات اخرى",
			"الاجمالي",
			"خصم التأمينات",
			"إضافات",
			"اجمالي الخصم",
			"صافى الراتب",
			"رقم الهوية",
		])
		worksheet.append([1, 2960, "موظف تجريبي", "محاسب", "الرياض", "الإدارة", "بنك", 5000, 1000, 500, 250, 6750, 500, 100, 900, 5950, "2089300780"])

		output = BytesIO()
		workbook.save(output)

		rows = payroll_module._extract_source_workbook_rows(output.getvalue())

		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0]["employee_id"], 2960)
		self.assertEqual(rows[0]["employee_name"], "موظف تجريبي")
		self.assertEqual(rows[0]["basic_salary"], 5000)
		self.assertEqual(rows[0]["gross_salary"], 6750)
		self.assertEqual(rows[0]["net_salary"], 5950)

	def test_extract_source_workbook_rows_prefers_payroll_print_sheet(self):
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.title = "كشف الرواتب طباعة"
		for _ in range(6):
			worksheet.append([None])
		worksheet.append([
			"التسلسل",
			"الرقم الوظيفي",
			"الاسم",
			"التحويل",
			"مكان العمل",
			"الإدارة",
			"بنك / كاش",
			"الاساسي",
			"بدل السكن",
			"بدل المواصلات",
			"بدلات اخرى",
			"الإضافي",
			"إجمالي البدلات",
			"ايام العمل",
			"ايام الغياب",
			"قيمة ايام الغياب",
			"ساعات التأخير",
			"قيمة التأخير",
			"خصم الجزاءات",
			"خصم التأمينات",
			"خصم السلف والاستقطاعات",
			"اجمالي الخصم",
			"صافى الراتب",
		])
		worksheet.append([
			1,
			2960,
			"موظف تجريبي",
			"شركة",
			"الرياض",
			"الإدارة",
			"بنك",
			5000,
			1000,
			500,
			250,
			100,
			6750,
			30,
			0,
			0,
			0,
			0,
			0,
			500,
			400,
			900,
			5950,
		])

		output = BytesIO()
		workbook.save(output)

		rows = payroll_module._extract_source_workbook_rows(output.getvalue())

		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0]["employee_id"], 2960)
		self.assertEqual(rows[0]["employee_name"], "موظف تجريبي")
		self.assertEqual(rows[0]["salary_mode"], "بنك")
		self.assertEqual(rows[0]["work_location"], "الرياض")
		self.assertEqual(rows[0]["additions"], 100)
		self.assertEqual(rows[0]["working_days"], 30)
		self.assertEqual(rows[0]["absence_days"], 0)
		self.assertEqual(rows[0]["late_hours"], 0)
		self.assertEqual(rows[0]["gross_salary"], 6750)
		self.assertEqual(rows[0]["gosi_deduction"], 500)
		self.assertEqual(rows[0]["total_deductions"], 900)
		self.assertEqual(rows[0]["net_salary"], 5950)

	def test_map_workbook_rows_to_payroll_preserves_other_deductions(self):
		raw_rows = [{
			"source_row": 6,
			"employee_id": 2960,
			"employee_name": "موظف تجريبي",
			"designation": "محاسب",
			"work_location": "الرياض",
			"salary_mode": "بنك",
			"basic_salary": 5000,
			"housing_allowance": 1000,
			"transport_allowance": 500,
			"other_allowances": 250,
			"gross_salary": 6750,
			"working_days": 30,
			"absence_days": 1,
			"late_hours": 2.5,
			"gosi_deduction": 500,
			"additions": 100,
			"total_deductions": 900,
			"net_salary": 5950,
			"national_id": "2089300780",
			"gosi_registration": "GOSI-1001",
		}]

		lookup = {
			"2960": {"employee": {"name": "EMP-2960", "employee_name": "موظف تجريبي", "department": "الإدارة", "nationality": "Saudi"}, "matched_by": "employee_id"}
		}

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value=lookup), patch.object(
			payroll_module.frappe.db, "exists", return_value=True
		):
			rows, warnings = payroll_module._map_workbook_rows_to_payroll("amd", raw_rows)

		self.assertEqual(len(rows), 1)
		self.assertEqual(warnings, [])
		self.assertEqual(rows[0]["employee"], "EMP-2960")
		self.assertEqual(rows[0]["designation"], "محاسب")
		self.assertEqual(rows[0]["work_location"], "الرياض")
		self.assertEqual(rows[0]["salary_mode"], "بنك")
		self.assertEqual(rows[0]["gosi_registration"], "GOSI-1001")
		self.assertEqual(rows[0]["working_days"], 30)
		self.assertEqual(rows[0]["absence_days"], 1)
		self.assertEqual(rows[0]["late_hours"], 2.5)
		self.assertEqual(rows[0]["gosi_employee_deduction"], 500)
		self.assertEqual(rows[0]["other_deductions"], 400)
		self.assertEqual(rows[0]["total_deductions"], 900)
		self.assertEqual(rows[0]["net_salary"], 5950)
		self.assertEqual(rows[0]["payroll_employee_id"], "2960")
		self.assertEqual(rows[0]["workbook_department"], "")

	def test_map_workbook_rows_to_payroll_normalizes_gross_when_it_includes_additions(self):
		raw_rows = [{
			"source_row": 15,
			"employee_id": 24841,
			"employee_name": "MUHAMMAD RIAZ",
			"basic_salary": 3030,
			"housing_allowance": 0,
			"transport_allowance": 0,
			"other_allowances": 1840,
			"gross_salary": 5981,
			"additions": 1111,
			"gosi_deduction": 0,
			"total_deductions": 0,
			"net_salary": 5981,
		}]

		lookup = {
			"24841": {"employee": {"name": "EMP-24841", "employee_name": "MUHAMMAD RIAZ", "department": "South", "nationality": "Saudi"}, "matched_by": "employee_id"}
		}

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value=lookup), patch.object(
			payroll_module.frappe.db, "exists", return_value=True
		):
			rows, warnings = payroll_module._map_workbook_rows_to_payroll("amd", raw_rows)

		self.assertEqual(len(rows), 1)
		self.assertEqual(warnings, [])
		self.assertEqual(rows[0]["gross_salary"], 4870)
		self.assertEqual(rows[0]["overtime_addition"], 1111)
		self.assertEqual(rows[0]["net_salary"], 5981)

	def test_validate_workbook_rows_accepts_gross_that_already_includes_additions(self):
		raw_rows = [{
			"source_row": 61,
			"employee_id": 20501,
			"employee_name": "FRANCIS RENARD ADRIANO",
			"basic_salary": 1997,
			"housing_allowance": 0,
			"transport_allowance": 0,
			"other_allowances": 544,
			"gross_salary": 3428,
			"additions": 887,
			"gosi_deduction": 0,
			"total_deductions": 0,
			"net_salary": 3428,
			"cost_center": "Central Region - Orbit - d",
		}]

		lookup = {
			"20501": {"employee": {"name": "20501", "employee_name": "FRANCIS RENARD ADRIANO"}, "matched_by": "employee_id"}
		}

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value=lookup), patch.object(
			payroll_module, "_get_postable_cost_center", return_value="Central Region - Orbit - d"
		):
			summary = payroll_module._validate_payroll_workbook_rows("ideaorbit company", raw_rows)

		self.assertEqual(summary["error_count"], 0)
		self.assertEqual(summary["errors"], [])

	def test_map_workbook_rows_to_payroll_keeps_unmatched_workbook_rows(self):
		raw_rows = [{
			"source_row": 8,
			"employee_id": "7803A",
			"employee_name": "MAJID ALI",
			"department": "الإدارة",
			"basic_salary": 750,
			"gross_salary": 750,
			"gosi_deduction": 0,
			"total_deductions": 0,
			"net_salary": 750,
		}]

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value={}), patch.object(
			payroll_module.frappe.db, "exists", return_value=False
		):
			rows, warnings = payroll_module._map_workbook_rows_to_payroll("amd", raw_rows)

		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0]["employee"], "")
		self.assertEqual(rows[0]["payroll_employee_id"], "7803A")
		self.assertEqual(rows[0]["employee_name"], "MAJID ALI")
		self.assertEqual(rows[0]["workbook_department"], "الإدارة")
		self.assertEqual(rows[0]["department"], "")
		self.assertIn("could not match employee 7803A", warnings[0])
		self.assertIn("without a linked Employee record", warnings[1])

	def test_match_workbook_employee_allows_trailing_letter_suffix(self):
		lookup = {
			"1090": {"employee": {"name": "EMP-1090", "employee_name": "Demo Employee"}, "matched_by": "employee_id"}
		}

		employee, matched_by = payroll_module._match_workbook_employee(
			{"employee_id": "1090A", "employee_name": "Demo Employee"},
			lookup,
		)

		self.assertEqual(employee["name"], "EMP-1090")
		self.assertEqual(matched_by, "employee_id")

	def test_preview_payroll_workbook_summarizes_matching_gap(self):
		with patch.object(payroll_module, "_extract_source_workbook_rows", return_value=[
			{"source_row": 6, "employee_id": 2960, "employee_name": "Missing Employee"}
		]), patch.object(payroll_module, "_get_attached_file_content", return_value=b"xlsx"), patch.object(payroll_module, "_map_workbook_rows_to_payroll", return_value=([{"employee_name": "Missing Employee"}], ["Row 6: could not match employee 2960."])), patch.object(
			payroll_module, "_collect_unmatched_workbook_rows", return_value=[{"source_row": 6, "employee_id": 2960, "employee_name": "Missing Employee"}]
		), patch.object(
			payroll_module.frappe.db, "count", return_value=1
		):
			summary = payroll_module._preview_payroll_workbook("amd", "/private/files/sample.xlsx")

		self.assertEqual(summary["total_rows"], 1)
		self.assertEqual(summary["importable_rows"], 1)
		self.assertEqual(summary["unmatched_rows"], 1)
		self.assertEqual(summary["company_employee_count"], 1)
		self.assertEqual(summary["sample_unmatched"], ["Row 6: could not match employee 2960."])
		self.assertEqual(summary["unmatched_details"][0]["employee_id"], 2960)

	def test_is_empty_source_row_skips_rows_without_employee_identity(self):
		self.assertTrue(payroll_module._is_empty_source_row({
			"employee_id": None,
			"employee_name": None,
			"gross_salary": 1500,
			"net_salary": 1200,
		}))

	def test_resolve_department_link_returns_only_existing_departments(self):
		with patch.object(payroll_module.frappe.db, "exists", side_effect=[True, False]):
			self.assertEqual(payroll_module._resolve_department_link("Finance"), "Finance")
			self.assertEqual(payroll_module._resolve_department_link("Unknown Department"), "")

	def test_build_basic_employee_payload_from_payroll_row_uses_defaults(self):
		row = SimpleNamespace(
			payroll_employee_id="7803A",
			employee_name="MAJID ALI",
			department="",
			workbook_department="Finance",
			designation="Analyst",
			nationality="Saudi",
		)
		defaults = {
			"gender": "Prefer not to say",
			"date_of_birth": getdate("1990-01-01"),
			"date_of_joining": getdate("2026-03-28"),
			"status": "Active",
		}

		with patch.object(payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: field in {"middle_name", "last_name", "department", "designation", "nationality"})), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True if (doctype == "Department" and name == "Finance") or (doctype == "Designation" and name == "Analyst") else False
		):
			payload = payroll_module._build_basic_employee_payload_from_payroll_row("amd", row, defaults)

		self.assertEqual(payload["employee_number"], "7803A")
		self.assertEqual(payload["first_name"], "MAJID")
		self.assertEqual(payload["last_name"], "ALI")
		self.assertEqual(payload["department"], "Finance")
		self.assertEqual(payload["designation"], "Analyst")
		self.assertEqual(payload["nationality"], "Saudi")
		self.assertEqual(payload["gender"], "Prefer not to say")

	def test_create_basic_employees_for_payroll_creates_and_links_rows(self):
		inserted = []

		class _EmployeeDoc:
			def __init__(self, payload):
				self.payload = payload
				self.flags = SimpleNamespace(ignore_permissions=False)
				self.name = f"EMP-{payload.get('employee_number') or payload['first_name']}"
				self.employee_name = " ".join(filter(None, [payload.get("first_name"), payload.get("middle_name"), payload.get("last_name")])).strip()
				self.department = payload.get("department")
				self.nationality = payload.get("nationality")
				self.employee_number = payload.get("employee_number")

			def insert(self):
				inserted.append(self.payload)
				return self

			def get(self, key):
				return getattr(self, key, self.payload.get(key))

		row_one = SimpleNamespace(idx=1, payroll_employee_id="7803A", employee="", employee_name="MAJID ALI", workbook_department="Finance", department="", nationality="")
		row_two = SimpleNamespace(idx=2, payroll_employee_id="7803A", employee="", employee_name="MAJID ALI", workbook_department="Finance", department="", nationality="")
		doc = SimpleNamespace(company="amd", employees=[row_one, row_two])
		defaults = {
			"gender": "Prefer not to say",
			"date_of_birth": getdate("1990-01-01"),
			"date_of_joining": getdate("2026-03-28"),
			"status": "Active",
		}

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value={}), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: field in {"middle_name", "last_name", "department"})
		), patch.object(
			payroll_module.frappe, "get_doc", side_effect=lambda payload: _EmployeeDoc(payload)
		), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True if doctype == "Department" and name == "Finance" else False
		):
			created, linked, skipped = payroll_module._create_basic_employees_for_payroll(doc, defaults)

		self.assertEqual(created, ["EMP-7803A"])
		self.assertEqual(linked, 1)
		self.assertEqual(skipped, [])
		self.assertEqual(len(inserted), 1)
		self.assertEqual(row_one.employee, "EMP-7803A")
		self.assertEqual(row_two.employee, "EMP-7803A")
		self.assertEqual(row_one.department, "Finance")

	def test_create_basic_employees_for_payroll_reuses_same_employee_for_suffix_variants(self):
		inserted = []

		class _EmployeeDoc:
			def __init__(self, payload):
				self.payload = payload
				self.flags = SimpleNamespace(ignore_permissions=False)
				self.name = f"EMP-{payload.get('employee_number') or payload['first_name']}"
				self.employee_name = " ".join(filter(None, [payload.get("first_name"), payload.get("middle_name"), payload.get("last_name")])).strip()
				self.department = payload.get("department")
				self.nationality = payload.get("nationality")
				self.employee_number = payload.get("employee_number")

			def insert(self):
				inserted.append(self.payload)
				return self

			def get(self, key):
				return getattr(self, key, self.payload.get(key))

		row_one = SimpleNamespace(idx=1, payroll_employee_id="7803", employee="", employee_name="MAJID ALI", workbook_department="Finance", department="", nationality="")
		row_two = SimpleNamespace(idx=2, payroll_employee_id="7803A", employee="", employee_name="MAJID ALI", workbook_department="Finance", department="", nationality="")
		doc = SimpleNamespace(company="amd", employees=[row_one, row_two])
		defaults = {
			"gender": "Prefer not to say",
			"date_of_birth": getdate("1990-01-01"),
			"date_of_joining": getdate("2026-03-28"),
			"status": "Active",
		}

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value={}), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: field in {"middle_name", "last_name", "department"})
		), patch.object(
			payroll_module.frappe, "get_doc", side_effect=lambda payload: _EmployeeDoc(payload)
		), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True if doctype == "Department" and name == "Finance" else False
		):
			created, linked, skipped = payroll_module._create_basic_employees_for_payroll(doc, defaults)

		self.assertEqual(created, ["EMP-7803"])
		self.assertEqual(linked, 1)
		self.assertEqual(skipped, [])
		self.assertEqual(len(inserted), 1)
		self.assertEqual(row_one.employee, "EMP-7803")
		self.assertEqual(row_two.employee, "EMP-7803")

	def test_create_basic_employees_for_payroll_keeps_same_name_with_unrelated_ids_separate(self):
		inserted = []

		class _EmployeeDoc:
			def __init__(self, payload):
				self.payload = payload
				self.flags = SimpleNamespace(ignore_permissions=False)
				self.name = f"EMP-{payload.get('employee_number') or payload['first_name']}"
				self.employee_name = " ".join(filter(None, [payload.get("first_name"), payload.get("middle_name"), payload.get("last_name")])).strip()
				self.department = payload.get("department")
				self.nationality = payload.get("nationality")
				self.employee_number = payload.get("employee_number")

			def insert(self):
				inserted.append(self.payload)
				return self

			def get(self, key):
				return getattr(self, key, self.payload.get(key))

		row_one = SimpleNamespace(idx=1, payroll_employee_id="7803", employee="", employee_name="MAJID ALI", workbook_department="Finance", department="", nationality="")
		row_two = SimpleNamespace(idx=2, payroll_employee_id="9811", employee="", employee_name="MAJID ALI", workbook_department="Finance", department="", nationality="")
		doc = SimpleNamespace(company="amd", employees=[row_one, row_two])
		defaults = {
			"gender": "Prefer not to say",
			"date_of_birth": getdate("1990-01-01"),
			"date_of_joining": getdate("2026-03-28"),
			"status": "Active",
		}

		with patch.object(payroll_module, "_get_company_employee_lookup", return_value={}), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: field in {"middle_name", "last_name", "department"})
		), patch.object(
			payroll_module.frappe, "get_doc", side_effect=lambda payload: _EmployeeDoc(payload)
		), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True if doctype == "Department" and name == "Finance" else False
		):
			created, linked, skipped = payroll_module._create_basic_employees_for_payroll(doc, defaults)

		self.assertEqual(created, ["EMP-7803", "EMP-9811"])
		self.assertEqual(linked, 0)
		self.assertEqual(skipped, [])
		self.assertEqual(len(inserted), 2)
		self.assertEqual(row_one.employee, "EMP-7803")
		self.assertEqual(row_two.employee, "EMP-9811")

	def test_clear_invalid_default_link_values_removes_invalid_custom_default(self):
		field = SimpleNamespace(fieldtype="Link", fieldname="salary_slip", options="Salary Slip", default="Salary Slip")

		class _Doc:
			doctype = "Employee"
			meta = SimpleNamespace(fields=[field])

			def __init__(self):
				self.salary_slip = "Salary Slip"

			def get(self, key):
				return getattr(self, key, None)

			def set(self, key, value):
				setattr(self, key, value)

		doc = _Doc()

		with patch.object(payroll_module.frappe.db, "exists", return_value=False):
			payroll_module._clear_invalid_default_link_values(doc, {"doctype": "Employee", "company": "amd"})

		self.assertEqual(doc.salary_slip, "")

	def test_apply_blank_overrides_for_invalid_default_links_sets_empty_payload_value(self):
		field = SimpleNamespace(fieldtype="Link", fieldname="salary_slip", options="Salary Slip", default="Salary Slip")
		payload = {"doctype": "Employee", "company": "amd"}

		with patch.object(payroll_module.frappe, "get_meta", return_value=SimpleNamespace(fields=[field])), patch.object(
			payroll_module.frappe.db, "exists", return_value=False
		):
			payroll_module._apply_blank_overrides_for_invalid_default_links("Employee", payload)

		self.assertEqual(payload["salary_slip"], "")

	def test_sync_linked_employee_master_fields_from_payroll_updates_blank_safe_fields(self):
		row = SimpleNamespace(
			employee="EMP-7803A",
			payroll_employee_id="7803A",
			workbook_department="Finance",
			department="",
			designation="Analyst",
			nationality="Saudi",
		)
		doc = SimpleNamespace(employees=[row])
		updates = []

		with patch.object(payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: field in {"designation", "nationality"})), patch.object(
			payroll_module.frappe.db, "get_value", return_value={
				"name": "EMP-7803A",
				"employee_number": "",
				"department": "",
				"designation": "",
				"nationality": "",
			}
		), patch.object(
			payroll_module.frappe.db, "set_value", side_effect=lambda doctype, name, values, update_modified=False: updates.append((doctype, name, values, update_modified))
		), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True if (doctype == "Department" and name == "Finance") or (doctype == "Designation" and name == "Analyst") else False
		):
			updated_count = payroll_module._sync_linked_employee_master_fields_from_payroll(doc, {"EMP-7803A"})

		self.assertEqual(updated_count, 1)
		self.assertEqual(updates[0][0], "Employee")
		self.assertEqual(updates[0][1], "EMP-7803A")
		self.assertEqual(updates[0][2], {
			"employee_number": "7803A",
			"department": "Finance",
			"designation": "Analyst",
			"nationality": "Saudi",
		})
		self.assertFalse(updates[0][3])
		self.assertEqual(row.department, "Finance")
		self.assertEqual(row.designation, "Analyst")
		self.assertEqual(row.nationality, "Saudi")

	def test_sync_linked_employee_master_fields_from_payroll_keeps_existing_values(self):
		row = SimpleNamespace(
			employee="EMP-7803A",
			payroll_employee_id="7803A",
			workbook_department="Finance",
			department="",
			designation="Analyst",
			nationality="Saudi",
		)
		doc = SimpleNamespace(employees=[row])

		with patch.object(payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: field in {"designation", "nationality"})), patch.object(
			payroll_module.frappe.db, "get_value", return_value={
				"name": "EMP-7803A",
				"employee_number": "EXISTING-1",
				"department": "Existing Department",
				"designation": "Existing Designation",
				"nationality": "Jordanian",
			}
		), patch.object(
			payroll_module.frappe.db, "set_value"
		) as set_value:
			updated_count = payroll_module._sync_linked_employee_master_fields_from_payroll(doc, {"EMP-7803A"})

		self.assertEqual(updated_count, 0)
		set_value.assert_not_called()
		self.assertEqual(row.department, "Existing Department")
		self.assertEqual(row.designation, "Existing Designation")
		self.assertEqual(row.nationality, "Jordanian")

	def test_delete_draft_payroll_uses_write_permission_and_deletes_doc(self):
		fake_doc = SimpleNamespace(name="SAU-PAY-TEST", docstatus=0, payroll_journal_entry="")

		with patch.object(payroll_module.frappe, "get_doc", return_value=fake_doc), patch.object(
			payroll_module.frappe, "has_permission"
		), patch.object(
			payroll_module.frappe.db, "count", return_value=0
		), patch.object(
			payroll_module.frappe, "delete_doc"
		) as delete_doc:
			result = payroll_module.delete_draft_payroll("SAU-PAY-TEST")

		delete_doc.assert_called_once_with("Saudi Monthly Payroll", "SAU-PAY-TEST", ignore_permissions=True)
		self.assertEqual(result, {"deleted": True, "name": "SAU-PAY-TEST"})

	def test_delete_draft_payroll_rejects_non_draft_docs(self):
		fake_doc = SimpleNamespace(name="SAU-PAY-TEST", docstatus=1, payroll_journal_entry="")

		with patch.object(payroll_module.frappe, "get_doc", return_value=fake_doc), patch.object(
			payroll_module.frappe, "has_permission"
		):
			with self.assertRaises(frappe.ValidationError):
				payroll_module.delete_draft_payroll("SAU-PAY-TEST")

	def test_normalize_basic_employee_creation_defaults_validates_gender_and_dates(self):
		with patch.object(payroll_module, "today", return_value="2026-04-02"), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True
		):
			defaults = payroll_module._normalize_basic_employee_creation_defaults(
				"Prefer not to say",
				"1990-01-01",
				"2026-03-28",
				"Active",
			)

		self.assertEqual(defaults["gender"], "Prefer not to say")
		self.assertEqual(str(defaults["date_of_birth"]), "1990-01-01")
		self.assertEqual(str(defaults["date_of_joining"]), "2026-03-28")

	def test_normalize_basic_employee_creation_defaults_rejects_future_dates(self):
		with patch.object(payroll_module, "today", return_value="2026-04-02"), patch.object(
			payroll_module.frappe.db, "exists", side_effect=lambda doctype, name=None: True
		):
			with self.assertRaises(frappe.ValidationError):
				payroll_module._normalize_basic_employee_creation_defaults(
					"Prefer not to say",
					"2030-01-01",
					"2030-02-01",
					"Active",
				)

	def test_get_payroll_payable_account_prefers_named_non_party_account(self):
		accounts = [
			{"name": "Creditors - A", "account_name": "Creditors", "account_type": "Payable"},
			{"name": "Payroll Payable - A", "account_name": "Payroll Payable", "account_type": ""},
		]

		with patch.object(payroll_module.frappe, "get_all", return_value=accounts):
			account = payroll_module._get_payroll_payable_account("amd")

		self.assertEqual(account, "Payroll Payable - A")

	def test_get_attached_file_content_rejects_invalid_extension(self):
		with patch.object(payroll_module.frappe.db, "get_value", return_value=frappe._dict({
			"name": "FILE-0001",
			"file_name": "payroll.csv",
			"file_size": 1024,
		})):
			with self.assertRaises(frappe.ValidationError):
				payroll_module._get_attached_file_content("/private/files/payroll.csv")

	def test_get_attached_file_content_rejects_large_file(self):
		with patch.object(payroll_module.frappe.db, "get_value", return_value=frappe._dict({
			"name": "FILE-0001",
			"file_name": "payroll.xlsx",
			"file_size": payroll_module.MAX_WORKBOOK_FILE_SIZE_BYTES + 1,
		})):
			with self.assertRaises(frappe.ValidationError):
				payroll_module._get_attached_file_content("/private/files/payroll.xlsx")

	def test_build_gap_report_rows_formats_unmatched_payload(self):
		rows = payroll_module._build_gap_report_rows([{
			"source_row": 6,
			"employee_id": 2960,
			"employee_name": "Missing Employee",
			"national_id": "2089300780",
			"department": "مدار الفكرة",
			"work_location": "الإدارة العامة",
			"designation": "مندوب مبيعات",
			"gross_salary": 7000,
			"net_salary": 7000,
		}])

		self.assertEqual(rows[0][0], "source_row")
		self.assertEqual(rows[1][1], 2960)
		self.assertEqual(rows[1][2], "Missing Employee")
		self.assertEqual(rows[1][9], "No matching Employee record in selected company")

	def test_employee_lookup_includes_related_saudi_identity_sources(self):
		employees = [{"name": "EMP-1", "employee_name": "Ali", "department": "HR", "employee_number": None}]
		contracts = [{"employee": "EMP-1", "iqama_number": "2089300780", "passport_number": "P123"}]
		permits = [{"employee": "EMP-1", "iqama_number": "2089300780"}]

		with patch.object(payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: False)), patch.object(
			payroll_module.frappe, "get_all", side_effect=[employees, contracts, permits]
		), patch.object(payroll_module.frappe.db, "exists", return_value=True):
			lookup = payroll_module._get_company_employee_lookup("amd")

		self.assertEqual(lookup["2089300780"]["employee"]["name"], "EMP-1")
		self.assertIn(lookup["2089300780"]["matched_by"], {"contract_iqama_number", "work_permit_iqama_number"})
		self.assertEqual(lookup["p123"]["employee"]["name"], "EMP-1")

	def test_build_employee_setup_template_prefills_gap_context(self):
		rows = payroll_module._build_employee_setup_template_rows("amd", [{
			"source_row": 6,
			"employee_id": 2960,
			"employee_name": "Missing Employee",
			"national_id": "2089300780",
			"department": "Finance",
			"designation": "Accountant",
		}])

		self.assertEqual(rows[0], payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS)
		self.assertEqual(rows[1][0], 6)
		self.assertEqual(rows[1][1], "2960")
		self.assertEqual(rows[1][4], "amd")
		self.assertEqual(rows[1][7], "2960")
		self.assertEqual(rows[1][8], "Missing")
		self.assertEqual(rows[1][9], "")
		self.assertEqual(rows[1][10], "Employee")
		self.assertEqual(rows[1][14], "Active")

	def test_split_payroll_employee_name_handles_multi_part_names(self):
		first_name, middle_name, last_name = payroll_module._split_payroll_employee_name("  Ahmad\nMohamed   Ali  Salem ")

		self.assertEqual(first_name, "Ahmad")
		self.assertEqual(middle_name, "Mohamed Ali")
		self.assertEqual(last_name, "Salem")

	def test_autofill_employee_setup_names_fills_missing_name_columns(self):
		rows, updated_count = payroll_module._autofill_employee_setup_names([{
			"payroll_employee_name": "Ahmad Mohamed Ali Salem",
			"first_name": "",
			"middle_name": None,
			"last_name": "",
		}])

		self.assertEqual(updated_count, 1)
		self.assertEqual(rows[0]["first_name"], "Ahmad")
		self.assertEqual(rows[0]["middle_name"], "Mohamed Ali")
		self.assertEqual(rows[0]["last_name"], "Salem")

	# ─── Template: verify the downloaded template leaves HR fields empty ─────────

	def test_employee_setup_template_has_empty_hr_fields_for_filling(self):
		"""القالب يجب أن يحتوي صف رأس صحيح، وأن تكون حقول HR (gender/dob/doj) فارغة."""
		rows = payroll_module._build_employee_setup_template_rows("amd", [{
			"source_row": 5,
			"employee_id": "9001",
			"employee_name": "Khalid Ibrahim Nasser",
			"national_id": "1000000001",
			"department": "Finance",
			"designation": "Analyst",
		}])

		# الصف الأول = رأس الأعمدة
		self.assertEqual(rows[0], payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS)

		data_row = rows[1]
		header = payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS

		gender_idx       = header.index("gender")
		dob_idx          = header.index("date_of_birth")
		doj_idx          = header.index("date_of_joining")
		first_name_idx   = header.index("first_name")
		last_name_idx    = header.index("last_name")

		# الحقول الجاهزة من ملف الرواتب يجب أن تكون ممتلئة
		self.assertEqual(data_row[first_name_idx], "Khalid")
		self.assertEqual(data_row[last_name_idx], "Nasser")

		# الحقول التي يجب أن يملأها HR يجب أن تكون فارغة
		self.assertEqual(data_row[gender_idx], "")
		self.assertEqual(data_row[dob_idx], "")
		self.assertEqual(data_row[doj_idx], "")

	def test_employee_setup_template_header_contains_all_expected_columns(self):
		"""تحقق من أن رأس القالب يحتوي جميع الأعمدة المطلوبة."""
		required = [
			"source_row", "payroll_employee_id", "payroll_employee_name",
			"national_id", "company", "department", "designation",
			"employee_number", "first_name", "middle_name", "last_name",
			"gender", "date_of_birth", "date_of_joining", "status", "remarks",
		]
		self.assertEqual(payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS, required)

	def test_employee_setup_template_empty_when_no_unmatched_rows(self):
		"""إذا لم توجد صفوف غير مطابقة، القالب يُعيد رأساً فقط بدون بيانات."""
		rows = payroll_module._build_employee_setup_template_rows("amd", [])
		self.assertEqual(len(rows), 1)  # header only
		self.assertEqual(rows[0], payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS)

	# ─── GOSI: تحقق من حساب الاقتطاع السعودي وغير السعودي ──────────────────────

	def test_build_employee_row_gosi_zero_for_non_saudi(self):
		"""الموظف غير السعودي لا يُقتطع منه GOSI."""
		emp = {"name": "EMP-0002", "employee_name": "John Doe", "department": "IT", "nationality": "American"}
		with patch.object(payroll_module, "get_employee_salary_components", return_value={
			"basic_salary": 5000, "housing_allowance": 0, "transport_allowance": 0, "other_allowances": 0,
		}), patch.object(payroll_module.frappe, "get_all", side_effect=[[], []]), \
		     patch.object(payroll_module, "get_gosi_rates", return_value={"employee_rate": 0.0}), \
		     patch.object(payroll_module, "get_due_loan_deduction", return_value={"loan_deduction": 0, "installment_names": []}):
			row = payroll_module._build_employee_row(emp, "March / مارس", 2026)

		self.assertEqual(row["gosi_employee_deduction"], 0.0)
		self.assertEqual(row["net_salary"], 5000.0)

	def test_build_employee_row_gosi_capped_at_45000_base(self):
		"""GOSI يُحسب على أساس 45,000 كحد أقصى وليس على الراتب الكامل."""
		emp = {"name": "EMP-0003", "employee_name": "Faisal", "department": "Exec", "nationality": "Saudi"}
		with patch.object(payroll_module, "get_employee_salary_components", return_value={
			"basic_salary": 60000, "housing_allowance": 0, "transport_allowance": 0, "other_allowances": 0,
		}), patch.object(payroll_module.frappe, "get_all", return_value=[]), \
		     patch.object(payroll_module, "get_gosi_rates", return_value={"employee_rate": 10.0}), \
		     patch.object(payroll_module, "get_due_loan_deduction", return_value={"loan_deduction": 0, "installment_names": []}):
			row = payroll_module._build_employee_row(emp, "March / مارس", 2026)

		expected_gosi = round(45000.0 * 10.0 / 100, 2)  # 4500.0
		self.assertEqual(row["gosi_employee_deduction"], expected_gosi)

	def test_build_employee_row_uses_settings_backed_gosi_rate(self):
		emp = {"name": "EMP-0004", "employee_name": "Demo Saudi", "department": "HR", "nationality": "Saudi"}
		with patch.object(payroll_module, "get_employee_salary_components", return_value={
			"basic_salary": 10000, "housing_allowance": 0, "transport_allowance": 0, "other_allowances": 0,
		}), patch.object(payroll_module.frappe, "get_all", side_effect=[[], []]), \
		     patch.object(payroll_module, "get_gosi_rates", return_value={"employee_rate": 7.5}), \
		     patch.object(payroll_module, "get_due_loan_deduction", return_value={"loan_deduction": 0, "installment_names": []}):
			row = payroll_module._build_employee_row(emp, "March / مارس", 2026)

		self.assertEqual(row["gosi_employee_deduction"], 750.0)

	# ─── صافي الراتب: التحقق من المعادلة الكاملة ────────────────────────────────

	def test_recalculate_employee_rows_net_equals_gross_plus_ot_minus_deductions(self):
		"""صافي الراتب = إجمالي + عمل إضافي - إجمالي الخصومات."""
		doc = frappe.get_doc({
			"doctype": "Saudi Monthly Payroll",
			"company": "amd",
			"month": "March / مارس",
			"year": 2026,
			"posting_date": "2026-03-31",
			"employees": [
				{
					"employee": "EMP-1",
					"basic_salary": 8000,
					"housing_allowance": 3000,
					"transport_allowance": 500,
					"other_allowances": 0,
					"gosi_employee_deduction": 800,
					"sick_leave_deduction": 100,
					"loan_deduction": 400,
					"other_deductions": 0,
					"overtime_addition": 250,
					"net_salary": 0,  # will be recalculated
				},
			],
		})
		doc._recalculate_employee_rows()

		row = doc.employees[0]
		self.assertEqual(row.gross_salary, 11500.0)
		self.assertEqual(row.total_deductions, 1300.0)
		self.assertEqual(row.net_salary, 10450.0)  # 11500 + 250 - 1300

	def test_validate_blocks_negative_net_salary(self):
		doc = frappe.get_doc({
			"doctype": "Saudi Monthly Payroll",
			"company": "amd",
			"month": "March / مارس",
			"year": 2026,
			"posting_date": "2026-03-31",
			"employees": [
				{
					"employee": "EMP-1",
					"employee_name": "Negative Case",
					"basic_salary": 1000,
					"housing_allowance": 0,
					"transport_allowance": 0,
					"other_allowances": 0,
					"gosi_employee_deduction": 100,
					"sick_leave_deduction": 200,
					"loan_deduction": 900,
					"other_deductions": 0,
					"overtime_addition": 0,
				}
			],
		})

		with self.assertRaises(frappe.ValidationError):
			doc.validate()

	def test_recalculate_totals_net_matches_sum_of_employee_nets(self):
		"""total_net_payable يجب أن يساوي مجموع صافي رواتب جميع الموظفين."""
		doc = frappe.get_doc({
			"doctype": "Saudi Monthly Payroll",
			"company": "amd",
			"month": "March / مارس",
			"year": 2026,
			"posting_date": "2026-03-31",
			"employees": [
				{"employee": "EMP-1", "gross_salary": 5000, "gosi_employee_deduction": 500, "sick_leave_deduction": 0, "loan_deduction": 0, "other_deductions": 0, "overtime_addition": 0, "net_salary": 4500},
				{"employee": "EMP-2", "gross_salary": 7500, "gosi_employee_deduction": 750, "sick_leave_deduction": 200, "loan_deduction": 300, "other_deductions": 50, "overtime_addition": 500, "net_salary": 6700},
			],
		})
		doc._recalculate_totals()

		self.assertEqual(doc.total_net_payable, 11200.0)
		self.assertEqual(doc.total_gross, 12500.0)
		self.assertEqual(doc.total_gosi_deductions, 1250.0)
		self.assertEqual(doc.total_sick_deductions, 200.0)
		self.assertEqual(doc.total_loan_deductions, 300.0)
		self.assertEqual(doc.total_other_deductions, 50.0)
		self.assertEqual(doc.total_overtime, 500.0)

	# ─── قالب الاستيراد: التحقق من إدارة الأسماء ────────────────────────────────

	def test_employee_setup_template_single_word_name(self):
		"""اسم من كلمة واحدة يُوضع في first_name فقط."""
		rows = payroll_module._build_employee_setup_template_rows("amd", [{
			"source_row": 1,
			"employee_id": "1001",
			"employee_name": "Mohammed",
			"national_id": "",
		}])
		header = payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS
		data = rows[1]
		self.assertEqual(data[header.index("first_name")], "Mohammed")
		self.assertEqual(data[header.index("middle_name")], "")
		self.assertEqual(data[header.index("last_name")], "")

	def test_employee_setup_template_two_word_name(self):
		"""اسم من كلمتين: first_name + last_name، middle_name فارغ."""
		rows = payroll_module._build_employee_setup_template_rows("amd", [{
			"source_row": 2,
			"employee_id": "1002",
			"employee_name": "Ali Hassan",
			"national_id": "",
		}])
		header = payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS
		data = rows[1]
		self.assertEqual(data[header.index("first_name")], "Ali")
		self.assertEqual(data[header.index("middle_name")], "")
		self.assertEqual(data[header.index("last_name")], "Hassan")

	# ─── is_empty_source_row: تحقق من منطق تخطي الصفوف الفارغة ─────────────────

	def test_is_empty_source_row_passes_rows_with_employee_id(self):
		"""صف يحتوي employee_id يجب ألّا يُعتبر فارغاً."""
		self.assertFalse(payroll_module._is_empty_source_row({
			"employee_id": "1023",
			"employee_name": None,
			"gross_salary": 0,
			"net_salary": 0,
		}))

	def test_is_empty_source_row_passes_rows_with_name_only(self):
		"""صف يحتوي اسم فقط بدون ID لا يُعتبر فارغاً."""
		self.assertFalse(payroll_module._is_empty_source_row({
			"employee_id": None,
			"employee_name": "Ahmad",
			"gross_salary": 0,
			"net_salary": 0,
		}))

	# ─── _get_payroll_payable_account: مزيد من حالات البحث عن الحساب ──────────

	def test_get_payroll_payable_account_prefers_salary_payable_over_party_payable(self):
		"""Salary Payable يُفضَّل على Creditors party account."""
		accounts = [
			{"name": "Creditors - A", "account_name": "Creditors", "account_type": "Payable"},
			{"name": "Salary Payable - A", "account_name": "Salary Payable", "account_type": ""},
		]
		with patch.object(payroll_module.frappe, "get_all", return_value=accounts):
			account = payroll_module._get_payroll_payable_account("amd")
		self.assertEqual(account, "Salary Payable - A")

	# ─────────────────────────────────────────────────────────────────────────────

	def test_extract_employee_setup_rows_reads_generated_workbook(self):
		with patch("frappe.utils.xlsxutils.get_excel_date_format", return_value=("yyyy-mm-dd", "hh:mm:ss")):
			content = payroll_module.make_xlsx([
				payroll_module.EMPLOYEE_SETUP_TEMPLATE_HEADERS,
				[6, "2960", "Missing Employee", "2089300780", "amd", "Finance", "Accountant", "2960", "Ali", "", "", "Male", "1990-01-01", "2024-01-01", "Active", "ok"],
			], "Employee Setup").getvalue()

		rows = payroll_module._extract_employee_setup_rows(content)

		self.assertEqual(len(rows), 1)
		self.assertEqual(rows[0]["employee_number"], "2960")
		self.assertEqual(rows[0]["first_name"], "Ali")

	def test_prepare_employee_setup_row_validates_required_fields(self):
		def _exists(doctype, name=None):
			if doctype == "Employee":
				return False
			if doctype == "DocType" and name == "Gender":
				return True
			if doctype in {"Gender", "Department", "Designation"}:
				return True
			return False

		with patch.object(payroll_module, "today", return_value="2026-04-02"), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: True)
		), patch.object(payroll_module.frappe.db, "exists", side_effect=_exists):
			payload = payroll_module._prepare_employee_setup_row("amd", {
				"source_row": 6,
				"employee_number": "2960",
				"payroll_employee_name": "Ali Ahmed Saleh",
				"company": "amd",
				"gender": "Male",
				"date_of_birth": "1990-01-01",
				"date_of_joining": "2024-01-01",
				"status": "Active",
				"department": "Finance",
				"designation": "Accountant",
			})

		self.assertEqual(payload["employee_number"], "2960")
		self.assertEqual(payload["first_name"], "Ali")
		self.assertEqual(payload["middle_name"], "Ahmed")
		self.assertEqual(payload["last_name"], "Saleh")
		self.assertEqual(payload["department"], "Finance")

	def test_prepare_employee_setup_row_rejects_invalid_dates_and_links(self):
		def _exists(doctype, name=None):
			if doctype == "Employee":
				return False
			if doctype == "DocType" and name == "Gender":
				return True
			if doctype == "Gender" and name == "Male":
				return True
			return False

		with patch.object(payroll_module, "today", return_value="2026-04-02"), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: True)
		), patch.object(payroll_module.frappe.db, "exists", side_effect=_exists):
			with self.assertRaises(frappe.ValidationError):
				payroll_module._prepare_employee_setup_row("amd", {
					"source_row": 6,
					"employee_number": "2960",
					"company": "amd",
					"first_name": "Ali",
					"gender": "Male",
					"date_of_birth": "2030-01-01",
					"date_of_joining": "2030-02-01",
					"status": "Active",
				})

		with patch.object(payroll_module, "today", return_value="2026-04-02"), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: True)
		), patch.object(payroll_module.frappe.db, "exists", side_effect=_exists):
			with self.assertRaises(frappe.ValidationError):
				payroll_module._prepare_employee_setup_row("amd", {
					"source_row": 7,
					"employee_number": "2961",
					"company": "amd",
					"first_name": "Omar",
					"gender": "Male",
					"date_of_birth": "1990-01-01",
					"date_of_joining": "2024-01-01",
					"department": "Unknown Department",
					"status": "Active",
				})

	def test_import_employee_setup_rows_creates_missing_employees_only(self):
		inserted = []

		class _EmployeeDoc:
			def __init__(self, payload):
				self.payload = payload
				self.flags = SimpleNamespace(ignore_permissions=False)
				self.name = f"EMP-{payload['employee_number']}"

			def insert(self):
				inserted.append(self.payload)
				return self

		def _exists(doctype, name=None):
			if doctype == "Employee":
				return name.get("employee_number") == "2961"
			if doctype == "DocType" and name == "Gender":
				return True
			if doctype == "Gender" and name == "Male":
				return True
			return False

		with patch.object(payroll_module, "today", return_value="2026-04-02"), patch.object(
			payroll_module.frappe, "get_meta", return_value=SimpleNamespace(has_field=lambda field: True)
		), patch.object(payroll_module.frappe.db, "exists", side_effect=_exists), patch.object(
			payroll_module.frappe, "get_doc", side_effect=lambda payload: _EmployeeDoc(payload)
		):
			created, skipped = payroll_module._import_employee_setup_rows("amd", [
				{
					"source_row": 6,
					"employee_number": "2960",
					"company": "amd",
					"first_name": "Ali",
					"gender": "Male",
					"date_of_birth": "1990-01-01",
					"date_of_joining": "2024-01-01",
					"status": "Active",
				},
				{
					"source_row": 7,
					"employee_number": "2961",
					"company": "amd",
					"first_name": "Omar",
					"gender": "Male",
					"date_of_birth": "1991-01-01",
					"date_of_joining": "2024-01-01",
					"status": "Active",
				},
			])

		self.assertEqual(created, ["EMP-2960"])
		self.assertEqual(len(inserted), 1)
		self.assertEqual(len(skipped), 1)