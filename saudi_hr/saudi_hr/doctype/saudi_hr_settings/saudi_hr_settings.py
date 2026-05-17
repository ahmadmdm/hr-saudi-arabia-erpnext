from __future__ import annotations

import json
from io import BytesIO
from os.path import splitext

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils.file_manager import save_file
from openpyxl import Workbook, load_workbook

from saudi_hr.saudi_hr.utils import assert_doctype_permissions


ALLOWED_IMPORT_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
MAX_IMPORT_FILE_SIZE_BYTES = 10 * 1024 * 1024


class SaudiHRSettings(Document):
	pass


def _assert_settings_access():
	frappe.only_for(("System Manager", "HR Manager"))


def _assert_attendance_api_admin():
	frappe.only_for(("System Manager",))


def _resolve_attendance_base_url(settings):
	from frappe.utils import get_url

	return (settings.mobile_attendance_base_url or get_url()).rstrip("/")


def _resource_endpoint(base_url: str, doctype: str):
	encoded_doctype = frappe.utils.quote(doctype)
	return {
		"doctype": doctype,
		"list": f"{base_url}/api/resource/{encoded_doctype}",
		"detail": f"{base_url}/api/resource/{encoded_doctype}/<name>",
		"create": f"POST {base_url}/api/resource/{encoded_doctype}",
		"update": f"PUT {base_url}/api/resource/{encoded_doctype}/<name>",
	}


def _build_attendance_api_settings_reference(settings):
	from saudi_hr.saudi_hr.api import MOBILE_ATTENDANCE_API_ENDPOINTS

	base_url = _resolve_attendance_base_url(settings)
	mobile_endpoints = []
	for endpoint in MOBILE_ATTENDANCE_API_ENDPOINTS:
		mobile_endpoints.append(
			{
				"key": endpoint["key"],
				"http_method": endpoint["http_method"],
				"method": endpoint["method"],
				"path": f"{base_url}/api/method/{endpoint['method']}",
				"description": endpoint["description"],
				"payload_fields": endpoint.get("payload_fields") or [],
			}
		)

	return {
		"title": "External HR Application API Reference",
		"base_url": base_url,
		"auth": {
			"scheme": "token",
			"header": "Authorization: token <api_key>:<api_secret>",
			"source": "Generate API Key and API Secret from the standard ERPNext User page. Saudi HR Settings does not generate or store credentials.",
			"example_curl_header": "-H 'Authorization: token YOUR_API_KEY:YOUR_API_SECRET'",
		},
		"admin_steps": [
			"Create a dedicated ERPNext User for the external app and assign only the roles it needs.",
			"Generate API Key and API Secret from the ERPNext User page, not from Saudi HR Settings.",
			"Set Base URL in Saudi HR Settings so the reference matches the target environment.",
			"Configure the external app to send Authorization: token <api_key>:<api_secret> on every request.",
			"Rotate the ERPNext User credentials from the User page if a device, server, or vendor changes.",
		],
		"mobile_attendance_methods": mobile_endpoints,
		"rest_resources": {
			"employee_profile": {
				"description": "Read employee master data, manager links, department, branch, company, status, and user mapping.",
				"resource": _resource_endpoint(base_url, "Employee"),
				"recommended_fields": ["name", "employee_name", "user_id", "company", "department", "branch", "reports_to", "status", "date_of_joining"],
			},
			"attendance_and_absence": {
				"description": "Create and read check-ins, approved attendance rows, absent days, and attendance locations.",
				"resources": [
					_resource_endpoint(base_url, "Employee Checkin"),
					_resource_endpoint(base_url, "Attendance"),
					_resource_endpoint(base_url, "Attendance Location"),
				],
				"absence_filter_example": f"{base_url}/api/resource/Attendance?filters=[[\"Attendance\",\"status\",\"=\",\"Absent\"]]",
			},
			"leave_and_requests": {
				"description": "Submit and track annual leave and ERPNext leave applications using the configured workflow.",
				"resources": [
					_resource_endpoint(base_url, "Saudi Annual Leave"),
					_resource_endpoint(base_url, "Leave Application"),
				],
			},
			"payroll_and_salary": {
				"description": "Read Saudi monthly payroll runs and ERPNext salary slips for employee-facing salary summaries.",
				"resources": [
					_resource_endpoint(base_url, "Saudi Monthly Payroll"),
					_resource_endpoint(base_url, "Saudi Monthly Payroll Employee"),
					_resource_endpoint(base_url, "Salary Slip"),
				],
				"privacy_note": "Expose salary endpoints only to a trusted integration user with tightly scoped roles and server-side filtering.",
			},
			"contracts_and_discipline": {
				"description": "Read employment contracts, warnings, investigations, and disciplinary procedures for HR self-service surfaces.",
				"resources": [
					_resource_endpoint(base_url, "Saudi Employment Contract"),
					_resource_endpoint(base_url, "Employee Warning Notice"),
					_resource_endpoint(base_url, "Investigation Record"),
					_resource_endpoint(base_url, "Disciplinary Procedure"),
				],
			},
			"organization_reference": {
				"description": "Reference data usually cached by external apps.",
				"resources": [
					_resource_endpoint(base_url, "Company"),
					_resource_endpoint(base_url, "Department"),
					_resource_endpoint(base_url, "Branch"),
					_resource_endpoint(base_url, "Designation"),
				],
			},
		},
		"request_examples": {
			"list_active_employees": f"GET {base_url}/api/resource/Employee?filters=[[\"Employee\",\"status\",\"=\",\"Active\"]]&fields=[\"name\",\"employee_name\",\"department\",\"branch\",\"reports_to\"]",
			"employee_salary_slips": f"GET {base_url}/api/resource/Salary%20Slip?filters=[[\"Salary Slip\",\"employee\",\"=\",\"<employee_id>\"],[\"Salary Slip\",\"docstatus\",\"=\",1]]",
			"employee_absences": f"GET {base_url}/api/resource/Attendance?filters=[[\"Attendance\",\"employee\",\"=\",\"<employee_id>\"],[\"Attendance\",\"status\",\"=\",\"Absent\"]]",
			"mobile_checkin": f"POST {base_url}/api/method/saudi_hr.saudi_hr.api.mobile_attendance_api_checkin",
		},
		"response_rules": {
			"success": "Frappe REST responses return data in the top-level data key. Custom mobile methods return {ok, data, error}.",
			"errors": "Handle HTTP 401/403 for invalid credentials or missing roles, 417 for validation errors, and 404 for missing records.",
			"dates": "Use ISO date strings: YYYY-MM-DD. Use site timezone for attendance timestamps unless the endpoint documents otherwise.",
		},
	}


def _get_employee_directory_rows():
	return frappe.get_all(
		"Employee",
		filters={"status": "Active"},
		fields=["name", "employee_name", "user_id", "branch", "department", "company"],
		order_by="employee_name asc, name asc",
	)


def _sync_directory_table():
	settings = frappe.get_single("Saudi HR Settings")
	assert_doctype_permissions("Saudi HR Settings", "write", doc=settings)
	settings.set("branch_employee_directory", [])
	for row in _get_employee_directory_rows():
		settings.append(
			"branch_employee_directory",
			{
				"employee": row.name,
				"employee_name": row.employee_name,
				"user_id": row.user_id,
				"branch": row.branch,
				"department": row.department,
				"company": row.company,
			},
		)
	settings.save()
	frappe.db.commit()
	return settings


def _save_attendance_api_reference(settings):
	settings.mobile_attendance_api_reference = json.dumps(
		_build_attendance_api_settings_reference(settings),
		indent=2,
		ensure_ascii=False,
	)
	settings.save()
	frappe.db.commit()
	return settings


def _build_template_file():
	workbook = Workbook()
	assignments = workbook.active
	assignments.title = "Employees"
	assignments.append(["employee_id", "employee_name", "user_id", "current_branch", "target_branch"])
	for row in _get_employee_directory_rows():
		assignments.append([row.name, row.employee_name, row.user_id, row.branch or "", row.branch or ""])

	branches_sheet = workbook.create_sheet("Branches")
	branches_sheet.append(["branch_name"])
	for branch_name in frappe.get_all("Branch", fields=["name"], order_by="name asc"):
		branches_sheet.append([branch_name.name])

	for sheet in workbook.worksheets:
		sheet.freeze_panes = "A2"

	buffer = BytesIO()
	workbook.save(buffer)
	buffer.seek(0)
	return buffer.getvalue()


def _normalize_header(value):
	return str(value or "").strip().lower().replace(" ", "_")


def _get_file_bytes(file_url: str) -> bytes:
	file_row = frappe.db.get_value("File", {"file_url": file_url}, ["name", "file_name", "file_size"], as_dict=True)
	if not file_row:
		frappe.throw(_("Unable to find the uploaded Excel file."))
	file_extension = splitext(str(file_row.file_name or file_url).strip())[1].lower()
	if file_extension not in ALLOWED_IMPORT_EXTENSIONS:
		frappe.throw(_("Only Excel workbook files are supported for branch imports."))
	if (file_row.file_size or 0) > MAX_IMPORT_FILE_SIZE_BYTES:
		frappe.throw(_("The uploaded Excel file is too large. Please keep it under 10 MB."))
	file_doc = frappe.get_doc("File", file_row.name)
	return file_doc.get_content()


def _ensure_branch(branch_name: str) -> str:
	branch_name = (branch_name or "").strip()
	if not branch_name:
		return ""
	if not frappe.db.exists("Branch", branch_name):
		assert_doctype_permissions("Branch", "create")
		frappe.get_doc({"doctype": "Branch", "branch": branch_name}).insert()
	return branch_name


def _find_employee_for_import(employee_id: str, user_id: str, employee_name: str):
	if employee_id and frappe.db.exists("Employee", employee_id):
		return employee_id
	if user_id:
		name = frappe.db.get_value("Employee", {"user_id": user_id}, "name")
		if name:
			return name
	if employee_name:
		name = frappe.db.get_value("Employee", {"employee_name": employee_name}, "name")
		if name:
			return name
	return None


@frappe.whitelist()
def sync_branch_employee_directory():
	_assert_settings_access()
	settings = _sync_directory_table()
	return {"row_count": len(settings.branch_employee_directory or [])}


@frappe.whitelist()
def download_employee_branch_template():
	_assert_settings_access()
	file_doc = save_file(
		"employee-branch-template.xlsx",
		_build_template_file(),
		"Saudi HR Settings",
		"Saudi HR Settings",
		is_private=0,
	)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name}


@frappe.whitelist()
def import_employee_branch_template(file_url: str | None = None):
	_assert_settings_access()
	if not file_url:
		frappe.throw(_("Attach the Excel template file first."))

	content = _get_file_bytes(file_url)
	workbook = load_workbook(BytesIO(content), data_only=True)
	worksheet = workbook[workbook.sheetnames[0]]
	rows = list(worksheet.iter_rows(values_only=True))
	if not rows:
		frappe.throw(_("The uploaded Excel file is empty."))

	headers = [_normalize_header(value) for value in rows[0]]
	updated_count = 0
	created_branch_count = 0
	skipped_count = 0
	errors = []

	for row in rows[1:]:
		payload = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
		employee_id = str(payload.get("employee_id") or "").strip()
		user_id = str(payload.get("user_id") or "").strip()
		employee_name = str(payload.get("employee_name") or "").strip()
		branch_name = str(payload.get("target_branch") or payload.get("branch") or payload.get("current_branch") or "").strip()

		if not any([employee_id, user_id, employee_name, branch_name]):
			continue
		if not branch_name:
			skipped_count += 1
			continue

		employee = _find_employee_for_import(employee_id, user_id, employee_name)
		if not employee:
			errors.append(_("Employee not found: {0}").format(employee_id or user_id or employee_name))
			continue

		if not frappe.db.exists("Branch", branch_name):
			_ensure_branch(branch_name)
			created_branch_count += 1

		current_branch = frappe.db.get_value("Employee", employee, "branch") or ""
		if current_branch == branch_name:
			skipped_count += 1
			continue

		frappe.db.set_value("Employee", employee, "branch", branch_name, update_modified=True)
		updated_count += 1

	frappe.db.commit()
	_sync_directory_table()
	return {
		"updated_count": updated_count,
		"created_branch_count": created_branch_count,
		"skipped_count": skipped_count,
		"errors": errors[:20],
	}


@frappe.whitelist()
def refresh_mobile_attendance_api_reference():
	_assert_attendance_api_admin()
	settings = frappe.get_single("Saudi HR Settings")
	_save_attendance_api_reference(settings)
	return {
		"base_url": _resolve_attendance_base_url(settings),
		"reference": settings.mobile_attendance_api_reference,
	}
