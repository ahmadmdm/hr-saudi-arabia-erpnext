from io import BytesIO
from unittest.mock import patch

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import load_workbook

from saudi_hr.saudi_hr.doctype.saudi_hr_settings import saudi_hr_settings as settings_module


class TestSaudiHRSettings(FrappeTestCase):
	def test_build_template_file_contains_employee_and_branch_sheets(self):
		employees = [
			frappe._dict({
				"name": "EMP-0001",
				"employee_name": "Administrator Mobile",
				"user_id": "administrator@example.com",
				"branch": "HQ",
			}),
		]
		branches = [
			frappe._dict({"name": "HQ"}),
			frappe._dict({"name": "Remote"}),
		]

		with patch.object(settings_module, "_get_employee_directory_rows", return_value=employees), patch.object(
			settings_module.frappe, "get_all", return_value=branches
		):
			content = settings_module._build_template_file()

		workbook = load_workbook(BytesIO(content), data_only=True)
		self.assertEqual(workbook.sheetnames, ["Employees", "Branches"])
		employee_sheet = workbook["Employees"]
		branch_sheet = workbook["Branches"]
		self.assertEqual(
			[cell.value for cell in employee_sheet[1]],
			["employee_id", "employee_name", "user_id", "current_branch", "target_branch"],
		)
		self.assertEqual(
			[cell.value for cell in employee_sheet[2]],
			["EMP-0001", "Administrator Mobile", "administrator@example.com", "HQ", "HQ"],
		)
		self.assertEqual(branch_sheet[2][0].value, "HQ")